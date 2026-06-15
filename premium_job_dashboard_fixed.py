# [file-tag: code-generated-file-1-1781426572390192427]
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
from collections import defaultdict
from typing import List, Dict, Any, Tuple

class JobHuntingAppGenerator:
    """
    Pythonを使用して「就活・インターン管理システム」を高度にデザインされたExcel形式で生成するクラス。
    単なるデータ出力ではなく、インタラクティブで堅牢なExcel製アプリを構築します。
    """
    def __init__(self, output_path: str = "就活管理システム.xlsx"):
        self.output_path = output_path
        self.wb = openpyxl.Workbook()
        
        # デフォルトの不要シートを削除
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])
            
        # 統一カラーパレット定義 (クラシック・ネイビー・モダンテーマ)
        self.NAVY_DARK = "1B365D"   # メインヘッダー
        self.NAVY_LIGHT = "F0F4F8"  # KPIカード値背景
        self.WHITE = "FFFFFF"       # 白文字
        self.GRAY_TEXT = "595959"   # サブテキスト
        self.GRAY_LIGHT = "F2F2F2"  # 偶数行縞模様用
        self.GRAY_BORDER = "D9D9D9" # 標準罫線
        
        # 優先度別の配色設定 (視認性の高いソフトトーン)
        self.PRIORITY_COLORS = {
            "S": {"fill": "FFD2D2", "font": "990000"}, # ソフト赤
            "A": {"fill": "FFE6CC", "font": "B35900"}, # ソフトオレンジ
            "B": {"fill": "D2E5FF", "font": "004C99"}, # ソフト青
            "C": {"fill": "E5E5E5", "font": "595959"}  # ソフトグレー
        }
        
        # イベントガントバー用の配色定義
        self.EVENT_COLORS = {
            "ES": {"fill": "FFF2CC", "font": "7F6000"},        # 黄色
            "SPI": {"fill": "E2EFDA", "font": "375623"},       # 緑
            "GD": {"fill": "E1D5E7", "font": "613A70"},        # 紫
            "面接": {"fill": "DDEBF7", "font": "1F4E78"},       # 青
            "インターン": {"fill": "FCE4D6", "font": "C65911"},   # オレンジ
            "内定": {"fill": "C9F2C9", "font": "1E461E"}        # 水色（ミント）
        }
        
        # ステータスマスター（ドロップダウン共通定義）
        self.STATUSES = [
            "応募済", "ES提出中", "ES通過", "ES不通過", "SPI受験", "SPI通過", "SPI不通過",
            "GD予定", "GD通過", "GD不通過", "一次面接通過", "二次面接通過", "最終面接通過",
            "結果待ち", "インターン参加予定", "インターン参加中", "インターン決定", "内定", "辞退"
        ]
        
        # モック・サンプルデータ
        self.sample_companies = [
            ["PwCコンサルティング", "コンサルティング", "S", "★★★★★", "面接中", "2026-05-10", "2026-06-05", "第一志望。面接対策必須"],
            ["三菱商事", "商社", "S", "★★★★★", "ES提出中", "2026-06-01", "2026-06-25", "OB訪問3名実施済"],
            ["トヨタ自動車", "メーカー", "A", "★★★★☆", "インターン参加予定", "2026-05-15", "2026-06-10", "インターン選考通過、8月参加"],
            ["リクルート", "IT・Web", "A", "★★★★☆", "SPI受験", "2026-05-20", "2026-06-15", "テストセンター受験予定"],
            ["野村総合研究所", "シンクタンク", "A", "★★★★☆", "GD予定", "2026-05-12", "2026-06-12", "グループディスカッション対策"],
            ["ソニー", "メーカー", "B", "★★★☆☆", "応募済", "2026-06-05", "2026-06-30", "本選考エントリー"],
            ["楽天グループ", "IT・Web", "B", "★★★☆☆", "内定", "2026-04-15", "2026-05-10", "早期内定。承諾保留中"],
            ["アクセンチュア", "コンサルティング", "B", "★★★☆☆", "ES通過", "2026-05-01", "2026-05-25", "次ステージ面接"],
            ["みずほフィナンシャルG", "金融", "C", "★★☆☆☆", "辞退", "2026-05-05", "2026-05-20", "他社内定のため辞退"]
        ]
        
        self.sample_events = [
            ["PwCコンサルティング", "ES", "2026-06-01", "2026-06-05", "志望動機のブラッシュアップ"],
            ["PwCコンサルティング", "面接", "2026-06-10", "2026-06-14", "二次面接（ケース含む）"],
            ["PwCコンサルティング", "インターン", "2026-06-22", "2026-06-26", "夏期5daysインターン"],
            ["三菱商事", "ES", "2026-06-15", "2026-06-25", "インターンエントリー"],
            ["トヨタ自動車", "面接", "2026-06-05", "2026-06-09", "インターン面接（パス済）"],
            ["リクルート", "SPI", "2026-06-12", "2026-06-18", "Webテスト受験期間"],
            ["野村総合研究所", "GD", "2026-06-18", "2026-06-21", "テーマ：新規事業立案"],
            ["ソニー", "ES", "2026-06-20", "2026-06-30", "技術職エントリーシート"]
        ]

    def generate_app(self) -> str:
        """すべてのシート群を正しい論理構造で構築し、Excelブックとして保存します。"""
        try:
            self._create_companies_sheet()
            self._create_events_sheet()
            self._create_dashboard_sheet()
            self._create_timeline_sheet()
            self._create_analytics_sheet()
            
            self.wb.save(self.output_path)
            return self.output_path
        except Exception as e:
            print(f"Error while generating the application: {str(e)}")
            raise e

    def _apply_base_grid(self, ws):
        """Excel標準のグリッド線が非表示にならないよう明示的に固定します。"""
        ws.views.sheetView[0].showGridLines = True

    def _create_dashboard_sheet(self):
        """Dashboardシートの作成: KPI、通過率表、直近スケジュール、埋め込みグラフ。"""
        ws = self.wb.create_sheet(title="Dashboard", index=0)
        self._apply_base_grid(ws)
        
        # タイネルバナー
        ws.merge_cells("A1:N1")
        title_cell = ws["A1"]
        title_cell.value = "  就活状況ダッシュボード (Dashboard)"
        title_cell.font = Font(name="Segoe UI", size=16, bold=True, color=self.WHITE)
        title_cell.fill = PatternFill(start_color=self.NAVY_DARK, end_color=self.NAVY_DARK, fill_type="solid")
        title_cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[1].height = 40
        
        # --- KPI カードセクション ---
        card_labels = [
            ("応募企業数", "=COUNTA(Companies!A2:A100)"),
            ("ES提出中", '=COUNTIF(Companies!E2:E100, "ES提出中")'),
            ("ES通過", '=COUNTIF(Companies!E2:E100, "ES通過")'),
            ("SPI受験", '=COUNTIF(Companies!E2:E100, "SPI受験")'),
            ("面接中", '=COUNTIF(Companies!E2:E100, "*面接*") + COUNTIF(Companies!E2:E100, "GD予定")'),
            ("結果待ち", '=COUNTIF(Companies!E2:E100, "結果待ち")'),
            ("インターン決定", '=COUNTIF(Companies!E2:E100, "インターン決定") + COUNTIF(Companies!E2:E100, "インターン参加予定")'),
            ("内定", '=COUNTIF(Companies!E2:E100, "内定")')
        ]
        
        thin_border = Border(left=Side(style='thin', color=self.GRAY_BORDER), right=Side(style='thin', color=self.GRAY_BORDER),
                             top=Side(style='thin', color=self.GRAY_BORDER), bottom=Side(style='thin', color=self.GRAY_BORDER))
        
        # 2行×4列のカード配置アドレス
        card_positions = [
            ("B3", "B4"), ("E3", "E4"), ("H3", "H4"), ("K3", "K4"),
            ("B6", "B7"), ("E6", "E7"), ("H6", "H7"), ("K6", "K7")
        ]
        
        for i, (label, formula) in enumerate(card_labels):
            pos_lbl, pos_val = card_positions[i]
            lbl_col, lbl_row = pos_lbl[0], int(pos_lbl[1:])
            val_col, val_row = pos_val[0], int(pos_val[1:])
            
            next_lbl_col = chr(ord(lbl_col) + 1)
            ws.merge_cells(f"{lbl_col}{lbl_row}:{next_lbl_col}{lbl_row}")
            ws.merge_cells(f"{val_col}{val_row}:{next_lbl_col}{val_row}")
            
            c_lbl = ws[f"{lbl_col}{lbl_row}"]
            c_lbl.value = label
            c_lbl.font = Font(name="Segoe UI", size=10, bold=True, color=self.WHITE)
            c_lbl.fill = PatternFill(start_color=self.NAVY_DARK, end_color=self.NAVY_DARK, fill_type="solid")
            c_lbl.alignment = Alignment(horizontal="center", vertical="center")
            
            c_val = ws[f"{val_col}{val_row}"]
            c_val.value = formula
            c_val.font = Font(name="Segoe UI", size=18, bold=True, color=self.NAVY_DARK)
            c_val.fill = PatternFill(start_color=self.NAVY_LIGHT, end_color=self.NAVY_LIGHT, fill_type="solid")
            c_val.alignment = Alignment(horizontal="center", vertical="center")
            
            # 結合セル領域全体に枠線を適用
            for r in range(lbl_row, val_row + 1):
                for c in [ord(lbl_col)-64, ord(lbl_col)-63]:
                    ws.cell(row=r, column=c).border = thin_border
                    
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[4].height = 30
        ws.row_dimensions[6].height = 20
        ws.row_dimensions[7].height = 30

        # --- 集計サブテーブルセクション (Row 10+) ---
        # 1. 選考通過率
        ws["B10"] = "選考通過率分析"
        ws["B10"].font = Font(name="Segoe UI", size=12, bold=True, color=self.NAVY_DARK)
        
        headers_rate = ["ステップ", "通過数", "通過率"]
        for col_idx, h in enumerate(headers_rate, start=2):
            cell = ws.cell(row=11, column=col_idx, value=h)
            cell.font = Font(name="Segoe UI", size=10, bold=True, color=self.WHITE)
            cell.fill = PatternFill(start_color="5A738E", end_color="5A738E", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            
        rates_data = [
            ("ES通過率", '=COUNTIF(Companies!E2:E100, "ES通過")+COUNTIF(Companies!E2:E100, "*面接*")+COUNTIF(Companies!E2:E100, "内定")+COUNTIF(Companies!E2:E100, "インターン*")', '=IF(B3>0, C12/B3, 0)'),
            ("SPI通過率", '=COUNTIF(Companies!E2:E100, "SPI通過")+COUNTIF(Companies!E2:E100, "*面接*")+COUNTIF(Companies!E2:E100, "内定")', '=IF(B3>0, C13/B3, 0)'),
            ("面接通過率", '=COUNTIF(Companies!E2:E100, "二次面接通過")+COUNTIF(Companies!E2:E100, "最終面接通過")+COUNTIF(Companies!E2:E100, "内定")', '=IF(B3>0, C14/B3, 0)'),
            ("インターン参加率", '=COUNTIF(Companies!E2:E100, "インターン参加中")+COUNTIF(Companies!E2:E100, "インターン決定")+COUNTIF(Companies!E2:E100, "インターン参加予定")', '=IF(B3>0, C15/B3, 0)')
        ]
        
        for r_idx, (step, form1, form2) in enumerate(rates_data, start=12):
            ws.cell(row=r_idx, column=2, value=step).font = Font(name="Segoe UI", size=10)
            c1 = ws.cell(row=r_idx, column=3, value=form1)
            c1.font = Font(name="Segoe UI", size=10)
            c1.alignment = Alignment(horizontal="right")
            
            c2 = ws.cell(row=r_idx, column=4, value=form2)
            c2.font = Font(name="Segoe UI", size=10, bold=True)
            c2.number_format = "0.0%"
            c2.alignment = Alignment(horizontal="right")
            for c in range(2, 5):
                ws.cell(row=r_idx, column=c).border = thin_border

        # 2. 優先度別集計テーブル
        ws["F10"] = "優先度別応募件数"
        ws["F10"].font = Font(name="Segoe UI", size=12, bold=True, color=self.NAVY_DARK)
        
        headers_pri = ["優先度", "件数"]
        for col_idx, h in enumerate(headers_pri, start=6):
            cell = ws.cell(row=11, column=col_idx, value=h)
            cell.font = Font(name="Segoe UI", size=10, bold=True, color=self.WHITE)
            cell.fill = PatternFill(start_color="5A738E", end_color="5A738E", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            
        pri_data = [
            ("Sランク", '=COUNTIF(Companies!C2:C100, "S")'),
            ("Aランク", '=COUNTIF(Companies!C2:C100, "A")'),
            ("Bランク", '=COUNTIF(Companies!C2:C100, "B")'),
            ("Cランク", '=COUNTIF(Companies!C2:C100, "C")'),
        ]
        for r_idx, (rank, form) in enumerate(pri_data, start=12):
            ws.cell(row=r_idx, column=6, value=rank).font = Font(name="Segoe UI", size=10)
            cc = ws.cell(row=r_idx, column=7, value=form)
            cc.font = Font(name="Segoe UI", size=10)
            cc.alignment = Alignment(horizontal="right")
            for c in range(6, 8):
                ws.cell(row=r_idx, column=c).border = thin_border

        # 3. 今週の予定 (直近のタスク一覧を綺麗に配置)
        ws["I10"] = "直近のイベント・選考予定"
        ws["I10"].font = Font(name="Segoe UI", size=12, bold=True, color=self.NAVY_DARK)
        
        headers_ev = ["企業名", "種別", "日付", "備考"]
        for col_idx, h in enumerate(headers_ev, start=9):
            cell = ws.cell(row=11, column=col_idx, value=h)
            cell.font = Font(name="Segoe UI", size=10, bold=True, color=self.WHITE)
            cell.fill = PatternFill(start_color="5A738E", end_color="5A738E", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            
        upcoming_sample = [
            ["PwCコンサルティング", "面接", "2026-06-10", "二次面接（ケース含む）"],
            ["リクルート", "SPI", "2026-06-12", "Webテスト受験期間"],
            ["三菱商事", "ES", "2026-06-15", "インターンエントリーシート締切"],
            ["野村総合研究所", "GD", "2026-06-18", "グループディスカッション対策"],
        ]
        
        for r_idx, row_data in enumerate(upcoming_sample, start=12):
            for col_idx, val in enumerate(row_data, start=9):
                cell = ws.cell(row=r_idx, column=col_idx, value=val)
                cell.font = Font(name="Segoe UI", size=10)
                cell.border = thin_border
                if col_idx == 11:
                    cell.alignment = Alignment(horizontal="center")

        # 優先度の分布を示す棒グラフをダッシュボード内に動的に生成埋め込み
        chart = BarChart()
        chart.type = "col"
        chart.title = "優先度別 企業数分布"
        chart.y_axis.title = "企業数"
        chart.x_axis.title = "優先度"
        data_ref = Reference(ws, min_col=7, min_row=11, max_row=15)
        cats_ref = Reference(ws, min_col=6, min_row=12, max_row=15)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.legend = None 
        chart.width = 11
        chart.height = 6.5
        ws.add_chart(chart, "B18")

        # 列幅自動調整
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            if col_letter in ['A', 'M', 'N']: continue
            ws.column_dimensions[col_letter].width = 14
        ws.column_dimensions['I'].width = 22
        ws.column_dimensions['L'].width = 24

    def _create_timeline_sheet(self):
        """Timelineシートの作成: 多段型タイムラインガントチャート、土日着色、今日ハイライト。"""
        ws = self.wb.create_sheet(title="Timeline", index=1)
        self._apply_base_grid(ws)
        
        # 上部凡例エリア
        ws["A1"] = "【凡例 (Legend)】"
        ws["A1"].font = Font(name="Segoe UI", size=10, bold=True)
        
        l_col = 2
        for ev_type, colors in self.EVENT_COLORS.items():
            cell = ws.cell(row=1, column=l_col, value=ev_type)
            cell.font = Font(name="Segoe UI", size=9, bold=True, color=colors["font"])
            cell.fill = PatternFill(start_color=colors["fill"], end_color=colors["fill"], fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(left=Side(style='thin', color=self.GRAY_BORDER), right=Side(style='thin', color=self.GRAY_BORDER))
            l_col += 1
            
        # 左側固定ヘッダー定義
        left_headers = ["企業名", "業界", "優先度", "志望度", "ステータス"]
        for i, h in enumerate(left_headers, start=1):
            ws.cell(row=3, column=i, value=h)
            ws.cell(row=3, column=i).font = Font(name="Segoe UI", size=10, bold=True, color=self.WHITE)
            ws.cell(row=3, column=i).fill = PatternFill(start_color=self.NAVY_DARK, end_color=self.NAVY_DARK, fill_type="solid")
            ws.cell(row=3, column=i).alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=3, start_column=i, end_row=4, end_column=i)
            
        # タイムライン日付設定 (2026年6月1日〜6月30日)
        start_date = datetime(2026, 6, 1)
        days_count = 30
        
        # 月ヘッダー (セル結合・濃紺背景)
        ws.merge_cells("F3:AI3")
        m_cell = ws["F3"]
        m_cell.value = "2026年 6月"
        m_cell.font = Font(name="Segoe UI", size=11, bold=True, color=self.WHITE)
        m_cell.fill = PatternFill(start_color="2E4A62", end_color="2E4A62", fill_type="solid")
        m_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        day_japanese = ["月", "火", "水", "木", "金", "土", "日"]
        
        # 日ヘッダー生成
        for d in range(days_count):
            cur_date = start_date + timedelta(days=d)
            col_idx = 6 + d
            
            cell_d = ws.cell(row=4, column=col_idx)
            cell_d.value = str(cur_date.day) + "\n(" + day_japanese[cur_date.weekday()] + ")"
            cell_d.font = Font(name="Segoe UI", size=8, color=self.WHITE)
            cell_d.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # 土日で色分け
            if cur_date.weekday() == 5: # 土
                cell_d.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            elif cur_date.weekday() == 6: # 日
                cell_d.fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
            else:
                cell_d.fill = PatternFill(start_color="415B76", end_color="415B76", fill_type="solid")
                
        ws.row_dimensions[3].height = 24
        ws.row_dimensions[4].height = 28
        
        # 左側固定列を維持するためウィンドウ枠を固定 (F5セル基準)
        ws.freeze_panes = "F5"
        
        # 重複選考を回避して多段サブ行にマッピングするためイベントを企業毎にグループ化
        events_by_company = defaultdict(list)
        for ev in self.sample_events:
            events_by_company[ev[0]].append({
                "type": ev[1],
                "start": datetime.strptime(ev[2], "%Y-%m-%d"),
                "end": datetime.strptime(ev[3], "%Y-%m-%d"),
                "note": ev[4]
            })
            
        current_row = 5
        thin_border = Border(left=Side(style='thin', color=self.GRAY_BORDER), right=Side(style='thin', color=self.GRAY_BORDER),
                             top=Side(style='thin', color=self.GRAY_BORDER), bottom=Side(style='thin', color=self.GRAY_BORDER))
        
        # ステータス列にドロップダウンバリデーションを付与
        dv_status = DataValidation(type="list", formula1="='Companies'!$I$2:$I$20", allow_blank=True)
        ws.add_data_validation(dv_status)
        
        for comp in self.sample_companies:
            c_name, industry, priority, star, status, _, _, _ = comp
            comp_events = events_by_company[c_name]
            
            # 重複判定を行いサブ行への割り当てを決定
            allocated_rows: List[List[Dict[str, Any]]] = []
            for ev in sorted(comp_events, key=lambda x: x["start"]):
                placed = False
                for row_list in allocated_rows:
                    if all(ev["start"] > e["end"] or ev["end"] < e["start"] for e in row_list):
                        row_list.append(ev)
                        placed = True
                        break
                if not placed:
                    allocated_rows.append([ev])
                    
            rows_needed = max(1, len(allocated_rows))
            
            # 左側固定列データの描画とマージ処理
            for col_idx, val in enumerate([c_name, industry, priority, star, status], start=1):
                cell_addr = f"{get_column_letter(col_idx)}{current_row}"
                target_cell = ws[cell_addr]
                target_cell.value = val
                target_cell.font = Font(name="Segoe UI", size=10)
                target_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                if rows_needed > 1:
                    end_cell_addr = f"{get_column_letter(col_idx)}{current_row + rows_needed - 1}"
                    ws.merge_cells(f"{cell_addr}:{end_cell_addr}")
                
                # 優先度別カラーハイライト
                if col_idx == 3:
                    cfg = self.PRIORITY_COLORS.get(priority, {"fill": "FFFFFF", "font": "000000"})
                    target_cell.fill = PatternFill(start_color=cfg["fill"], end_color=cfg["fill"], fill_type="solid")
                    target_cell.font = Font(name="Segoe UI", size=10, bold=True, color=cfg["font"])
                
                if col_idx == 5:
                    dv_status.add(target_cell)
            
            # 右側タイムライン背景グリッドおよび土日・今日の強調
            for r_offset in range(rows_needed):
                r_idx = current_row + r_offset
                ws.row_dimensions[r_idx].height = 22
                
                for d in range(days_count):
                    c_date = start_date + timedelta(days=d)
                    g_col = 6 + d
                    grid_cell = ws.cell(row=r_idx, column=g_col)
                    grid_cell.border = thin_border
                    
                    if c_date.weekday() == 5:
                        grid_cell.fill = PatternFill(start_color="F2F6FB", end_color="F2F6FB", fill_type="solid")
                    elif c_date.weekday() == 6:
                        grid_cell.fill = PatternFill(start_color="FDF2F2", end_color="FDF2F2", fill_type="solid")
                        
                    # 「今日(2026/06/14)」の強調赤枠
                    if c_date.date() == datetime(2026, 6, 14).date():
                        grid_cell.border = Border(
                            left=Side(style='medium', color="FF0000"), right=Side(style='medium', color="FF0000"),
                            top=Side(style='thin', color=self.GRAY_BORDER), bottom=Side(style='thin', color=self.GRAY_BORDER)
                        )
            
            # ガントチャートバー（選考イベント）を重ねて配置
            for r_offset, row_list in enumerate(allocated_rows):
                r_idx = current_row + r_offset
                for ev in row_list:
                    s_offset = (ev["start"] - start_date).days
                    e_offset = (ev["end"] - start_date).days
                    
                    s_col = max(0, s_offset)
                    e_col = min(days_count - 1, e_offset)
                    
                    start_col_idx = 6 + s_col
                    end_col_idx = 6 + e_col
                    
                    if start_col_idx <= end_col_idx:
                        ws.merge_cells(start_row=r_idx, start_column=start_col_idx, end_row=r_idx, end_column=end_col_idx)
                        bar_cell = ws.cell(row=r_idx, column=start_col_idx, value=ev["type"])
                        
                        cfg = self.EVENT_COLORS.get(ev["type"], {"fill": "E5E5E5", "font": "000000"})
                        bar_cell.fill = PatternFill(start_color=cfg["fill"], end_color=cfg["fill"], fill_type="solid")
                        bar_cell.font = Font(name="Segoe UI", size=9, bold=True, color=cfg["font"])
                        bar_cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        for bc in range(start_col_idx, end_col_idx + 1):
                            ws.cell(row=r_idx, column=bc).border = thin_border
            
            for r_offset in range(rows_needed):
                for c_idx in range(1, 6):
                    ws.cell(row=current_row + r_offset, column=c_idx).border = thin_border
                    
            current_row += rows_needed
            
        # カラム幅のチューニング
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 18
        for d in range(days_count):
            ws.column_dimensions[get_column_letter(6 + d)].width = 6

    def _create_companies_sheet(self):
        """Companiesシートの作成: 企業情報のマスターDBデータ。"""
        ws = self.wb.create_sheet(title="Companies", index=2)
        self._apply_base_grid(ws)
        
        headers = ["企業名", "業界", "優先度", "志望度", "ステータス", "応募日", "ES締切", "備考"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(name="Segoe UI", size=11, bold=True, color=self.WHITE)
            cell.fill = PatternFill(start_color=self.NAVY_DARK, end_color=self.NAVY_DARK, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        ws.row_dimensions[1].height = 26
        thin_border = Border(left=Side(style='thin', color=self.GRAY_BORDER), right=Side(style='thin', color=self.GRAY_BORDER),
                             top=Side(style='thin', color=self.GRAY_BORDER), bottom=Side(style='thin', color=self.GRAY_BORDER))
                             
        for r_idx, row_data in enumerate(self.sample_companies, start=2):
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = Font(name="Segoe UI", size=10)
                cell.border = thin_border
                if c_idx in [3, 4, 5, 6, 7]:
                    cell.alignment = Alignment(horizontal="center")
            ws.row_dimensions[r_idx].height = 20

        # マスタ選択肢定義用の隠し・参照用カラム
        ws.cell(row=1, column=9, value="マスタ選択肢").font = Font(bold=True)
        for idx, st in enumerate(self.STATUSES, start=2):
            ws.cell(row=idx, column=9, value=st)
            
        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 30

    def _create_events_sheet(self):
        """Eventsシートの作成: 各企業の選考スケジュール期間。"""
        ws = self.wb.create_sheet(title="Events", index=3)
        self._apply_base_grid(ws)
        
        headers = ["企業名", "イベント種別", "開始日", "終了日", "備考"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(name="Segoe UI", size=11, bold=True, color=self.WHITE)
            cell.fill = PatternFill(start_color=self.NAVY_DARK, end_color=self.NAVY_DARK, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        ws.row_dimensions[1].height = 26
        thin_border = Border(left=Side(style='thin', color=self.GRAY_BORDER), right=Side(style='thin', color=self.GRAY_BORDER),
                             top=Side(style='thin', color=self.GRAY_BORDER), bottom=Side(style='thin', color=self.GRAY_BORDER))
                             
        for r_idx, row_data in enumerate(self.sample_events, start=2):
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = Font(name="Segoe UI", size=10)
                cell.border = thin_border
                if c_idx in [2, 3, 4]:
                    cell.alignment = Alignment(horizontal="center")
            ws.row_dimensions[r_idx].height = 20
            
        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 32

    def _create_analytics_sheet(self):
        """Analyticsシートの作成: 円グラフ（ステータス分布）と折れ線グラフ（応募推移）。"""
        ws = self.wb.create_sheet(title="Analytics", index=4)
        self._apply_base_grid(ws)
        
        # タイトルバナー
        ws.merge_cells("A1:K1")
        title_cell = ws["A1"]
        title_cell.value = "  就職活動 選考・統計分析レポーティング (Analytics)"
        title_cell.font = Font(name="Segoe UI", size=14, bold=True, color=self.WHITE)
        title_cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        title_cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[1].height = 35

        # 1. ステータス分布集計データ
        ws["A3"] = "ステータス分布データ"
        ws["A3"].font = Font(name="Segoe UI", bold=True)
        ws["A4"] = "ステータス"
        ws["B4"] = "社数"
        ws["A4"].fill = PatternFill(start_color=self.NAVY_LIGHT, end_color=self.NAVY_LIGHT, fill_type="solid")
        ws["B4"].fill = PatternFill(start_color=self.NAVY_LIGHT, end_color=self.NAVY_LIGHT, fill_type="solid")
        
        analytics_statuses = ["内定", "面接中", "ES提出中", "応募済", "インターン決定", "辞退"]
        formulas_statuses = [
            '=COUNTIF(Companies!E2:E100, "内定")',
            '=COUNTIF(Companies!E2:E100, "*面接*") + COUNTIF(Companies!E2:E100, "GD予定")',
            '=COUNTIF(Companies!E2:E100, "ES提出中")',
            '=COUNTIF(Companies!E2:E100, "応募済")',
            '=COUNTIF(Companies!E2:E100, "インターン決定") + COUNTIF(Companies!E2:E100, "インターン参加予定")',
            '=COUNTIF(Companies!E2:E100, "辞退")'
        ]
        
        thin_border = Border(left=Side(style='thin', color=self.GRAY_BORDER), right=Side(style='thin', color=self.GRAY_BORDER),
                             top=Side(style='thin', color=self.GRAY_BORDER), bottom=Side(style='thin', color=self.GRAY_BORDER))
                             
        for idx, (st, form) in enumerate(zip(analytics_statuses, formulas_statuses), start=5):
            ws.cell(row=idx, column=1, value=st).font = Font(name="Segoe UI", size=9)
            c2 = ws.cell(row=idx, column=2, value=form)
            c2.font = Font(name="Segoe UI", size=9)
            c2.alignment = Alignment(horizontal="right")
            ws.cell(row=idx, column=1).border = thin_border
            ws.cell(row=idx, column=2).border = thin_border

        # 円グラフ埋め込み
        pie = PieChart()
        pie.title = "選考ステータス比率"
        labels = Reference(ws, min_col=1, min_row=5, max_row=10)
        data = Reference(ws, min_col=2, min_row=4, max_row=10)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.style = 10
        pie.width = 12
        pie.height = 7.5
        ws.add_chart(pie, "D3")

        # 2. 月別応募推移
        ws["A13"] = "月別応募推移"
        ws["A13"].font = Font(name="Segoe UI", bold=True)
        ws["A14"] = "月"
        ws["B14"] = "応募数"
        
        months = ["2026年4月", "2026年5月", "2026年6月"]
        m_formulas = [
            '=COUNTIF(Companies!F2:F100, "*2026-04*")',
            '=COUNTIF(Companies!F2:F100, "*2026-05*")',
            '=COUNTIF(Companies!F2:F100, "*2026-06*")'
        ]
        for idx, (m, f) in enumerate(zip(months, m_formulas), start=15):
            ws.cell(row=idx, column=1, value=m).font = Font(name="Segoe UI", size=9)
            c2 = ws.cell(row=idx, column=2, value=f)
            c2.font = Font(name="Segoe UI", size=9)
            c2.alignment = Alignment(horizontal="right")
            ws.cell(row=idx, column=1).border = thin_border
            ws.cell(row=idx, column=2).border = thin_border

        # 折れ線グラフ埋め込み
        line = LineChart()
        line.title = "選考エントリー数推移"
        line.y_axis.title = "社数"
        line.x_axis.title = "時期"
        l_data = Reference(ws, min_col=2, min_row=14, max_row=17)
        l_cats = Reference(ws, min_col=1, min_row=15, max_row=17)
        line.add_data(l_data, titles_from_data=True)
        line.set_categories(l_cats)
        line.legend = None
        line.width = 12
        line.height = 6.5
        ws.add_chart(line, "D14")

        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 10

if __name__ == "__main__":
    generator = JobHuntingAppGenerator()
    saved_file = generator.generate_app()
    print(f"Application executed successfully! Saved workbook as: {saved_file}")