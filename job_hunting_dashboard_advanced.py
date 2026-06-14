
"""
Advanced Job Hunting Dashboard
pip install openpyxl
"""

from datetime import datetime, timedelta
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

COMPANIES = [
    ["PwC","コンサル","S",5,"面接中"],
    ["NTTデータ","IT","A",4,"ES通過"],
    ["アクセンチュア","コンサル","S",5,"結果待ち"],
]

EVENTS = [
    ["PwC","ES","2026-06-20","2026-06-23"],
    ["PwC","面接","2026-06-28","2026-06-28"],
    ["PwC","インターン","2026-08-05","2026-08-07"],
    ["NTTデータ","SPI","2026-06-22","2026-06-23"],
    ["アクセンチュア","面接","2026-07-10","2026-07-10"],
]

STATUS_LIST = [
    "応募済","ES提出中","ES通過","ES不通過","SPI受験",
    "面接予定","面接中","結果待ち","インターン決定","内定","辞退"
]

COLORS = {
    "ES":"FFD966",
    "SPI":"A9D18E",
    "面接":"9DC3E6",
    "インターン":"F4B183"
}

wb = Workbook()

# Dashboard
dash = wb.active
dash.title = "Dashboard"

# Companies
company_ws = wb.create_sheet("Companies")
company_ws.append(["企業","業界","優先度","志望度(1-5)","ステータス"])
for r in COMPANIES:
    company_ws.append(r)

# dropdowns
dv = DataValidation(type="list", formula1='"' + ",".join(STATUS_LIST) + '"')
company_ws.add_data_validation(dv)
for r in range(2, company_ws.max_row+1):
    dv.add(company_ws[f"E{r}"])

# Events
event_ws = wb.create_sheet("Events")
event_ws.append(["企業","イベント","開始","終了"])
for e in EVENTS:
    event_ws.append(e)

# Timeline
timeline = wb.create_sheet("Timeline")

dates = []
for _,_,s,e in EVENTS:
    cur = datetime.strptime(s,"%Y-%m-%d").date()
    end = datetime.strptime(e,"%Y-%m-%d").date()
    while cur <= end:
        dates.append(cur)
        cur += timedelta(days=1)

start=min(dates)
end=max(dates)

timeline["A1"]="企業"
timeline["B1"]="状態"

companies=[c[0] for c in COMPANIES]
status_map={c[0]:c[4] for c in COMPANIES}

col=3
date_map={}
cur=start
while cur<=end:
    timeline.cell(2,col,cur.day)
    timeline.cell(3,col,"月火水木金土日"[cur.weekday()])
    date_map[cur]=col
    cur+=timedelta(days=1)
    col+=1

for i,c in enumerate(companies,start=4):
    timeline.cell(i,1,c)
    timeline.cell(i,2,status_map[c])

for company,event,s,e in EVENTS:
    row=companies.index(company)+4
    sc=date_map[datetime.strptime(s,"%Y-%m-%d").date()]
    ec=date_map[datetime.strptime(e,"%Y-%m-%d").date()]

    if ec>sc:
        timeline.merge_cells(start_row=row,start_column=sc,end_row=row,end_column=ec)

    timeline.cell(row,sc,event)
    timeline.cell(row,sc).alignment=Alignment(horizontal="center")

    fill=PatternFill("solid",fgColor=COLORS.get(event,"D9D9D9"))
    for c in range(sc,ec+1):
        timeline.cell(row,c).fill=fill

today=datetime.today().date()
if start<=today<=end:
    tc=date_map[today]
    for r in range(1,timeline.max_row+1):
        timeline.cell(r,tc).fill=PatternFill("solid",fgColor="FFCCCC")

for d,c in date_map.items():
    if d.weekday()>=5:
        for r in range(4,timeline.max_row+1):
            if timeline.cell(r,c).value is None:
                timeline.cell(r,c).fill=PatternFill("solid",fgColor="EEEEEE")

# Analytics
ana=wb.create_sheet("Analytics")

statuses=[c[4] for c in COMPANIES]
cnt=Counter(statuses)

ana.append(["状態","数"])
for k,v in cnt.items():
    ana.append([k,v])

pie=PieChart()
labels=Reference(ana,min_col=1,min_row=2,max_row=ana.max_row)
data=Reference(ana,min_col=2,min_row=1,max_row=ana.max_row)
pie.add_data(data,titles_from_data=True)
pie.set_categories(labels)
pie.title="ステータス分布"
ana.add_chart(pie,"E2")

ana["A10"]="応募数"
ana["B10"]=len(COMPANIES)

# Dashboard content
dash["A1"]="就活ダッシュボード"
dash["A1"].font=Font(size=18,bold=True)

dash["A3"]="応募企業数"
dash["B3"]=len(COMPANIES)

dash["A4"]="S優先度"
dash["B4"]=sum(1 for c in COMPANIES if c[2]=="S")

dash["A6"]="企業一覧は Companies"
dash["A7"]="工程表は Timeline"
dash["A8"]="分析は Analytics"

# styling
for ws in wb.worksheets:
    for col in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(col)].width = 15

wb.active = 0
wb.save("job_hunting_dashboard_advanced.xlsx")
print("created")
