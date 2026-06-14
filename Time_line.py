from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
    Border,
    Side
)
from openpyxl.utils import get_column_letter

# ====================================
# データ入力
# ====================================

events = [

    {
        "company": "A社",
        "status": "面接中",
        "event": "ES",
        "start": "2026-06-20",
        "end": "2026-06-23"
    },

    {
        "company": "A社",
        "status": "面接中",
        "event": "面接",
        "start": "2026-06-25",
        "end": "2026-06-25"
    },

    {
        "company": "A社",
        "status": "インターン決定",
        "event": "インターン",
        "start": "2026-08-05",
        "end": "2026-08-07"
    },

    {
        "company": "A社",
        "status": "インターン決定",
        "event": "インターン",
        "start": "2026-08-12",
        "end": "2026-08-12"
    },

    {
        "company": "A社",
        "status": "インターン決定",
        "event": "インターン",
        "start": "2026-08-19",
        "end": "2026-08-19"
    },

    {
        "company": "B社",
        "status": "SPI受験",
        "event": "SPI",
        "start": "2026-06-21",
        "end": "2026-06-22"
    },

    {
        "company": "B社",
        "status": "結果待ち",
        "event": "面接",
        "start": "2026-06-28",
        "end": "2026-06-28"
    }
]

# ====================================
# 色設定
# ====================================

COLORS = {
    "ES": "FFD966",
    "SPI": "A9D18E",
    "面接": "9DC3E6",
    "インターン": "F4B183",
    "合格": "C6E0B4"
}

HEADER_COLOR = "203864"
TODAY_COLOR = "FFF2CC"
WEEKEND_COLOR = "F2F2F2"

# ====================================
# 日付範囲
# ====================================

all_dates = []

for e in events:
    start = datetime.strptime(e["start"], "%Y-%m-%d")
    end = datetime.strptime(e["end"], "%Y-%m-%d")

    current = start

    while current <= end:
        all_dates.append(current)
        current += timedelta(days=1)

start_date = min(all_dates)
end_date = max(all_dates)

# ====================================
# Workbook
# ====================================

wb = Workbook()
ws = wb.active
ws.title = "Job Dashboard"

# ====================================
# Header
# ====================================

ws["A1"] = "企業"
ws["B1"] = "ステータス"

header_fill = PatternFill(
    "solid",
    fgColor=HEADER_COLOR
)

header_font = Font(
    color="FFFFFF",
    bold=True
)

for cell in ["A1", "B1"]:
    ws[cell].fill = header_fill
    ws[cell].font = header_font

date_col = 3

date_map = {}

current = start_date

today = datetime.today().date()

while current <= end_date:

    cell = ws.cell(
        row=1,
        column=date_col
    )

    cell.value = current.strftime("%m/%d")

    cell.fill = header_fill
    cell.font = header_font

    date_map[current.date()] = date_col

    date_col += 1
    current += timedelta(days=1)

# ====================================
# Company Row
# ====================================

companies = sorted(
    list(
        set(
            e["company"]
            for e in events
        )
    )
)

row_map = {}

for idx, company in enumerate(
    companies,
    start=2
):
    row_map[company] = idx

    ws.cell(
        row=idx,
        column=1
    ).value = company

# status
for company in companies:

    statuses = [
        e["status"]
        for e in events
        if e["company"] == company
    ]

    ws.cell(
        row=row_map[company],
        column=2
    ).value = statuses[-1]

# ====================================
# 土日・今日
# ====================================

current = start_date

while current <= end_date:

    col = date_map[current.date()]

    if current.weekday() >= 5:

        for r in range(
            2,
            len(companies)+2
        ):
            ws.cell(
                r,
                col
            ).fill = PatternFill(
                "solid",
                fgColor=WEEKEND_COLOR
            )

    if current.date() == today:

        for r in range(
            1,
            len(companies)+2
        ):
            ws.cell(
                r,
                col
            ).fill = PatternFill(
                "solid",
                fgColor=TODAY_COLOR
            )

    current += timedelta(days=1)

# ====================================
# Event Plot
# ====================================

for event in events:

    row = row_map[event["company"]]

    start = datetime.strptime(
        event["start"],
        "%Y-%m-%d"
    )

    end = datetime.strptime(
        event["end"],
        "%Y-%m-%d"
    )

    current = start

    while current <= end:

        col = date_map[current.date()]

        cell = ws.cell(
            row=row,
            column=col
        )

        cell.value = event["event"]

        fill_color = COLORS.get(
            event["event"],
            "D9D9D9"
        )

        cell.fill = PatternFill(
            "solid",
            fgColor=fill_color
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        current += timedelta(days=1)

# ====================================
# Styling
# ====================================

thin = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

for row in ws.iter_rows():

    for cell in row:
        cell.border = thin

# 列幅

ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 18

for col in range(
    3,
    ws.max_column + 1
):
    ws.column_dimensions[
        get_column_letter(col)
    ].width = 5

# 固定表示

ws.freeze_panes = "C2"

# フィルタ

ws.auto_filter.ref = ws.dimensions

# ====================================
# 保存
# ====================================

wb.save("job_hunting_dashboard.xlsx")

print("job_hunting_dashboard.xlsx を作成しました")