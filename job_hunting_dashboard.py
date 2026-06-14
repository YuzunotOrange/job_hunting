"""
Job Hunting Dashboard Generator
Requirements:
    pip install openpyxl

Run:
    python job_hunting_dashboard.py

Edit the EVENTS list below and rerun.
"""

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

STATUS_LIST = [
    "応募済", "ES提出中", "ES通過", "ES不通過",
    "SPI受験", "SPI通過", "SPI不通過",
    "GD予定", "GD通過", "GD不通過",
    "面接予定", "面接中", "結果待ち",
    "インターン参加予定", "インターン参加中",
    "インターン決定", "内定", "辞退"
]

EVENTS = [
    {"company":"A社","status":"インターン決定","event":"ES","start":"2026-06-20","end":"2026-06-23"},
    {"company":"A社","status":"インターン決定","event":"面接","start":"2026-06-28","end":"2026-06-28"},
    {"company":"A社","status":"インターン決定","event":"インターン","start":"2026-08-05","end":"2026-08-07"},
    {"company":"A社","status":"インターン決定","event":"インターン","start":"2026-08-19","end":"2026-08-19"},
    {"company":"B社","status":"結果待ち","event":"SPI","start":"2026-06-22","end":"2026-06-23"},
    {"company":"B社","status":"結果待ち","event":"面接","start":"2026-07-10","end":"2026-07-10"},
]

COLORS = {
    "ES": "FFD966",
    "SPI": "A9D18E",
    "GD": "D9D2E9",
    "面接": "9DC3E6",
    "インターン": "F4B183",
}

wb = Workbook()
ws = wb.active
ws.title = "工程表"

dates = []
for e in EVENTS:
    s = datetime.strptime(e["start"], "%Y-%m-%d")
    t = datetime.strptime(e["end"], "%Y-%m-%d")
    while s <= t:
        dates.append(s.date())
        s += timedelta(days=1)

start_date = min(dates)
end_date = max(dates)

companies = sorted({e["company"] for e in EVENTS})

ws["A1"] = "企業名"
ws["B1"] = "ステータス"

header_fill = PatternFill("solid", fgColor="203864")
header_font = Font(color="FFFFFF", bold=True)

for c in ("A1", "B1"):
    ws[c].fill = header_fill
    ws[c].font = header_font

# 3-row calendar header
date_map = {}
col = 3
cur = start_date

while cur <= end_date:
    ws.cell(1, col, cur.strftime("%Y/%m"))
    ws.cell(2, col, cur.day)
    ws.cell(3, col, "月火水木金土日"[cur.weekday()])
    date_map[cur] = col
    col += 1
    cur += timedelta(days=1)

# merge month headers
start_col = 3
current_month = None
month_start = 3
for c in range(3, col):
    m = ws.cell(1, c).value
    if current_month is None:
        current_month = m
    elif m != current_month:
        ws.merge_cells(start_row=1, start_column=month_start,
                       end_row=1, end_column=c-1)
        current_month = m
        month_start = c
ws.merge_cells(start_row=1, start_column=month_start,
               end_row=1, end_column=col-1)

for r in range(1, 4):
    for c in range(3, col):
        ws.cell(r, c).fill = header_fill
        ws.cell(r, c).font = header_font
        ws.cell(r, c).alignment = Alignment(horizontal="center")

row_map = {}
for i, company in enumerate(companies, start=4):
    row_map[company] = i
    ws.cell(i, 1, company)

    status = next(e["status"] for e in EVENTS if e["company"] == company)
    ws.cell(i, 2, status)

# dropdown status
dv = DataValidation(type="list", formula1='"' + ",".join(STATUS_LIST) + '"')
ws.add_data_validation(dv)
for r in range(4, 4 + len(companies)):
    dv.add(ws.cell(r, 2))

# bars
for e in EVENTS:
    row = row_map[e["company"]]
    s = datetime.strptime(e["start"], "%Y-%m-%d").date()
    t = datetime.strptime(e["end"], "%Y-%m-%d").date()

    start_col = date_map[s]
    end_col = date_map[t]

    if start_col != end_col:
        ws.merge_cells(start_row=row, start_column=start_col,
                       end_row=row, end_column=end_col)

    cell = ws.cell(row, start_col)
    cell.value = e["event"]
    cell.alignment = Alignment(horizontal="center", vertical="center")

    fill = PatternFill("solid", fgColor=COLORS.get(e["event"], "D9D9D9"))

    for c in range(start_col, end_col + 1):
        ws.cell(row, c).fill = fill

# widths
ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 20
for c in range(3, col):
    ws.column_dimensions[get_column_letter(c)].width = 4

# borders
thin = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))

for row in ws.iter_rows():
    for cell in row:
        cell.border = thin

ws.freeze_panes = "C4"
ws.auto_filter.ref = ws.dimensions

wb.save("job_hunting_dashboard.xlsx")
print("Created job_hunting_dashboard.xlsx")
