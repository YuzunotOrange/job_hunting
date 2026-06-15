# job_hunting_system_v2.py
# -*- coding: utf-8 -*-

"""
就活管理システム v2 Pro
--------------------------------------------
初回のみ実行

生成ファイル:
    就活管理システム_v2.xlsx

設計方針:
    ・Pythonは新規作成のみ
    ・更新処理なし
    ・load_workbook禁止
    ・Excel365側で運用
"""

import os
from datetime import datetime

from openpyxl import Workbook

from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)

from openpyxl.utils import get_column_letter

from openpyxl.worksheet.table import (
    Table,
    TableStyleInfo
)

from openpyxl.worksheet.datavalidation import (
    DataValidation
)

from openpyxl.chart import (
    BarChart,
    PieChart,
    Reference
)

from openpyxl.formatting.rule import FormulaRule


# ==========================================================
# 基本設定
# ==========================================================

OUTPUT_FILE = "就活管理システム_v2.xlsx"

VERSION = "v2 Pro"

CREATED_AT = datetime.now().strftime("%Y-%m-%d")


# ==========================================================
# ファイル存在チェック
# ==========================================================

if os.path.exists(OUTPUT_FILE):
    raise SystemExit(
        f"エラー: {OUTPUT_FILE} は既に存在します。\n"
        f"上書きは禁止されています。"
    )


# ==========================================================
# Workbook生成
# ==========================================================

wb = Workbook()


# ==========================================================
# シート生成
# ==========================================================

ws_dashboard = wb.active
ws_dashboard.title = "Dashboard"

ws_companies = wb.create_sheet("Companies")
ws_events = wb.create_sheet("Events")
ws_timeline = wb.create_sheet("Timeline")
ws_analytics = wb.create_sheet("Analytics")
ws_settings = wb.create_sheet("Settings")
ws_master = wb.create_sheet("_Master")


# ==========================================================
# 共通スタイル
# ==========================================================

NAVY = "1F3A5F"
LIGHT_GRAY = "F3F4F6"
BORDER_GRAY = "D0D7DE"

thin_side = Side(
    border_style="thin",
    color=BORDER_GRAY
)

card_border = Border(
    left=thin_side,
    right=thin_side,
    top=thin_side,
    bottom=thin_side
)

header_fill = PatternFill(
    fill_type="solid",
    fgColor=NAVY
)

card_fill = PatternFill(
    fill_type="solid",
    fgColor="FFFFFF"
)

gray_fill = PatternFill(
    fill_type="solid",
    fgColor=LIGHT_GRAY
)

header_font = Font(
    color="FFFFFF",
    bold=True,
    size=11
)

title_font = Font(
    bold=True,
    size=14
)

normal_font = Font(
    size=10
)

card_value_font = Font(
    bold=True,
    size=18
)

center_alignment = Alignment(
    horizontal="center",
    vertical="center"
)

left_alignment = Alignment(
    horizontal="left",
    vertical="center"
)


# ==========================================================
# 列幅設定用ヘルパー
# ==========================================================

def set_width(ws, widths):
    """
    widths:
        {
            "A":20,
            "B":30
        }
    """
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# ==========================================================
# _Master シート
# ==========================================================

MASTER_INDUSTRIES = [
    "コンサル",
    "商社",
    "メーカー",
    "IT・Web",
    "シンクタンク",
    "金融",
    "人材",
    "通信",
    "不動産",
    "広告",
    "インフラ",
    "物流",
    "公務員",
    "その他"
]

MASTER_PRIORITY = [
    "S",
    "A",
    "B",
    "C"
]

MASTER_MOTIVATION = [
    "★★★★★",
    "★★★★☆",
    "★★★☆☆",
    "★★☆☆☆",
    "★☆☆☆☆"
]

MASTER_STATUS = [
    "応募予定",
    "応募済",

    "ES提出中",
    "ES通過",
    "ES不通過",

    "SPI受験",
    "SPI通過",
    "SPI不通過",

    "GD予定",
    "GD通過",
    "GD不通過",

    "一次面接予定",
    "一次面接通過",

    "二次面接予定",
    "二次面接通過",

    "最終面接予定",
    "最終面接通過",

    "結果待ち",

    "インターン参加予定",
    "インターン参加中",
    "インターン決定",

    "内定",

    "辞退"
]

MASTER_EVENTS = [
    "説明会",
    "ES",
    "SPI",
    "GD",
    "一次面接",
    "二次面接",
    "最終面接",
    "インターン",
    "内定"
]


# ヘッダ

ws_master["A1"] = "業界"
ws_master["B1"] = "優先度"
ws_master["C1"] = "志望度"
ws_master["D1"] = "ステータス"
ws_master["E1"] = "イベント"

for cell in ws_master[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = card_border


# 業界

for row, value in enumerate(MASTER_INDUSTRIES, start=2):
    ws_master.cell(row=row, column=1, value=value)

# 優先度

for row, value in enumerate(MASTER_PRIORITY, start=2):
    ws_master.cell(row=row, column=2, value=value)

# 志望度

for row, value in enumerate(MASTER_MOTIVATION, start=2):
    ws_master.cell(row=row, column=3, value=value)

# ステータス

for row, value in enumerate(MASTER_STATUS, start=2):
    ws_master.cell(row=row, column=4, value=value)

# イベント

for row, value in enumerate(MASTER_EVENTS, start=2):
    ws_master.cell(row=row, column=5, value=value)


set_width(
    ws_master,
    {
        "A": 20,
        "B": 15,
        "C": 15,
        "D": 25,
        "E": 20
    }
)


# ==========================================================
# Settings
# ==========================================================

ws_settings["A1"] = "Version"
ws_settings["A2"] = "作成日"
ws_settings["A3"] = "最終更新日"

ws_settings["B1"] = VERSION
ws_settings["B2"] = CREATED_AT

# Excel側で常に今日
ws_settings["B3"] = "=TODAY()"

for row in range(1, 4):
    ws_settings[f"A{row}"].font = title_font

set_width(
    ws_settings,
    {
        "A": 20,
        "B": 25
    }
)


# ==========================================================
# シート共通初期設定
# ==========================================================

for ws in [
    ws_dashboard,
    ws_companies,
    ws_events,
    ws_timeline,
    ws_analytics,
    ws_settings
]:
    ws.sheet_view.showGridLines = True


# ==========================================================
# _Master 非表示
# ==========================================================

ws_master.sheet_state = "hidden"


# ==========================================================
# Part1終了
# ==========================================================
# 保存処理は最終Partで実施
# wb.save()はまだ呼ばない

# ==========================================================
# Part2 v2.1
# Companiesシート（実運用向け）
# ==========================================================

COMPANY_HEADERS = [
    "企業名",
    "業界",
    "優先度",
    "志望度",
    "ステータス",
    "応募日",
    "ES締切",
    "締切警告",
    "マイページURL",
    "備考"
]

# ----------------------------------------------------------
# ヘッダー
# ----------------------------------------------------------

for col_num, header in enumerate(COMPANY_HEADERS, start=1):

    cell = ws_companies.cell(
        row=1,
        column=col_num,
        value=header
    )

    cell.fill = header_fill
    cell.font = header_font
    cell.border = card_border
    cell.alignment = center_alignment

# ----------------------------------------------------------
# 列幅
# ----------------------------------------------------------

set_width(
    ws_companies,
    {
        "A": 32,
        "B": 18,
        "C": 10,
        "D": 12,
        "E": 20,
        "F": 14,
        "G": 14,
        "H": 14,
        "I": 45,
        "J": 40
    }
)

# ----------------------------------------------------------
# テーブル初期行
#
# Excel Tableは最低1データ行が必要
# ----------------------------------------------------------

ws_companies["A2"] = ""

# ----------------------------------------------------------
# 締切警告列
#
# Structured Reference採用
# Excelが自動継承
# ----------------------------------------------------------

ws_companies["H2"] = (
    '=IF([@[ES締切]]="",'
    ' "",'
    ' IF([@[ES締切]]<TODAY(),"🔴期限切れ",'
    ' IF([@[ES締切]]=TODAY(),"🔴今日",'
    ' IF([@[ES締切]]<=TODAY()+3,"🟠3日以内",'
    ' IF([@[ES締切]]<=TODAY()+7,"🟡7日以内","")))))'
)

# ----------------------------------------------------------
# テーブル作成
# ----------------------------------------------------------

company_table = Table(
    displayName="tblCompanies",
    ref="A1:J2"
)

company_style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)

company_table.tableStyleInfo = company_style

ws_companies.add_table(company_table)

# ----------------------------------------------------------
# 入力規則
#
# すべて_Master参照
# ----------------------------------------------------------

dv_industry = DataValidation(
    type="list",
    formula1="=_Master!$A$2:$A$15",
    allow_blank=True
)

dv_priority = DataValidation(
    type="list",
    formula1="=_Master!$B$2:$B$5",
    allow_blank=True
)

dv_motivation = DataValidation(
    type="list",
    formula1="=_Master!$C$2:$C$6",
    allow_blank=True
)

dv_status = DataValidation(
    type="list",
    formula1="=_Master!$D$2:$D$24",
    allow_blank=True
)

for dv in (
    dv_industry,
    dv_priority,
    dv_motivation,
    dv_status
):
    ws_companies.add_data_validation(dv)

# ----------------------------------------------------------
# 将来行まで適用
#
# テーブルが伸びても有効
# ----------------------------------------------------------

dv_industry.add("B2:B1048576")
dv_priority.add("C2:C1048576")
dv_motivation.add("D2:D1048576")
dv_status.add("E2:E1048576")

# ----------------------------------------------------------
# 日付列
# ----------------------------------------------------------

ws_companies["F2"].number_format = "yyyy/mm/dd"
ws_companies["G2"].number_format = "yyyy/mm/dd"

# ----------------------------------------------------------
# URL列
# ----------------------------------------------------------

ws_companies["I2"].style = "Hyperlink"

# ----------------------------------------------------------
# フリーズ
# ----------------------------------------------------------

ws_companies.freeze_panes = "A2"

# ----------------------------------------------------------
# オートフィルタ
# ----------------------------------------------------------

ws_companies.auto_filter.ref = "A1:J2"

# ----------------------------------------------------------
# ユーザーガイド
# ----------------------------------------------------------

guide_col = "L"

ws_companies[f"{guide_col}1"] = "入力ガイド"

ws_companies[f"{guide_col}2"] = "企業名を入力"

ws_companies[f"{guide_col}3"] = "業界はプルダウン"

ws_companies[f"{guide_col}4"] = "優先度はS/A/B/C"

ws_companies[f"{guide_col}5"] = "志望度は★評価"

ws_companies[f"{guide_col}6"] = "ステータス選択"

ws_companies[f"{guide_col}7"] = "ES締切で警告自動判定"

ws_companies[f"{guide_col}8"] = "URL保存可能"

for row in range(1, 9):

    c = ws_companies[f"{guide_col}{row}"]

    c.border = card_border

    if row == 1:
        c.fill = header_fill
        c.font = header_font

# ----------------------------------------------------------
# 将来のDashboard/Timeline向け
#
# Companiesシート内の補助セル
# （非表示列ではない）
# ----------------------------------------------------------

ws_companies["N1"] = "Companies_Count"

ws_companies["N2"] = '=COUNTA(tblCompanies[企業名])'

# ----------------------------------------------------------
# 印刷時見やすさ
# ----------------------------------------------------------

ws_companies.sheet_view.showGridLines = True

# ==========================================================
# Part2 v2.1 終了
# ==========================================================
# ==========================================================
# Part3 v2.1
# Eventsシート
# ==========================================================

EVENT_HEADERS = [
    "企業名",
    "イベント",
    "開始日",
    "終了日",
    "開始時刻",
    "備考"
]

# ----------------------------------------------------------
# ヘッダ
# ----------------------------------------------------------

for col_num, header in enumerate(EVENT_HEADERS, start=1):

    cell = ws_events.cell(
        row=1,
        column=col_num,
        value=header
    )

    cell.fill = header_fill
    cell.font = header_font
    cell.border = card_border
    cell.alignment = center_alignment


# ----------------------------------------------------------
# 列幅
# ----------------------------------------------------------

set_width(
    ws_events,
    {
        "A": 30,
        "B": 18,
        "C": 14,
        "D": 14,
        "E": 12,
        "F": 40
    }
)


# ----------------------------------------------------------
# テーブル初期行
#
# Excel Tableは最低1行必要
# ----------------------------------------------------------

ws_events["A2"] = ""


# ----------------------------------------------------------
# 日付列書式
# ----------------------------------------------------------

ws_events["C2"].number_format = "yyyy/mm/dd"

ws_events["D2"].number_format = "yyyy/mm/dd"


# ----------------------------------------------------------
# 時刻列書式
# ----------------------------------------------------------

ws_events["E2"].number_format = "hh:mm"


# ----------------------------------------------------------
# Excelテーブル
# ----------------------------------------------------------

events_table = Table(
    displayName="tblEvents",
    ref="A1:F2"
)

events_style = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)

events_table.tableStyleInfo = events_style

ws_events.add_table(events_table)


# ----------------------------------------------------------
# イベント入力規則
#
# _Master参照
# ----------------------------------------------------------

dv_event = DataValidation(
    type="list",
    formula1="=_Master!$E$2:$E$10",
    allow_blank=True
)

ws_events.add_data_validation(dv_event)

dv_event.add("B2:B1048576")


# ----------------------------------------------------------
# 企業名候補
#
# Companiesから選択可能にする準備
#
# ※Part7でNamedRange版へ移行予定
# ----------------------------------------------------------

# 現時点では自由入力


# ----------------------------------------------------------
# フリーズ
# ----------------------------------------------------------

ws_events.freeze_panes = "A2"


# ----------------------------------------------------------
# オートフィルタ
# ----------------------------------------------------------

ws_events.auto_filter.ref = "A1:F2"


# ----------------------------------------------------------
# ユーザーガイド
# ----------------------------------------------------------

guide_col = "H"

ws_events[f"{guide_col}1"] = "入力ガイド"

ws_events[f"{guide_col}2"] = "企業名を入力"

ws_events[f"{guide_col}3"] = "イベントはプルダウン"

ws_events[f"{guide_col}4"] = "開始日を入力"

ws_events[f"{guide_col}5"] = "終了日を入力"

ws_events[f"{guide_col}6"] = "開始時刻を入力"

ws_events[f"{guide_col}7"] = "Timelineへ自動反映"

ws_events[f"{guide_col}8"] = "Dashboardへ自動反映"

for row in range(1, 9):

    c = ws_events[f"{guide_col}{row}"]

    c.border = card_border

    if row == 1:

        c.fill = header_fill
        c.font = header_font


# ----------------------------------------------------------
# 補助セル
#
# Dashboard集計用
# ----------------------------------------------------------

ws_events["J1"] = "Events_Count"

ws_events["J2"] = '=COUNTA(tblEvents[企業名])'


# ----------------------------------------------------------
# 本日のイベント数
# ----------------------------------------------------------

ws_events["J4"] = "Today_Events"

ws_events["J5"] = (
    '=COUNTIF(tblEvents[開始日],TODAY())'
)


# ----------------------------------------------------------
# 未来イベント数
# ----------------------------------------------------------

ws_events["J7"] = "Future_Events"

ws_events["J8"] = (
    '=COUNTIF(tblEvents[開始日],">="&TODAY())'
)


# ----------------------------------------------------------
# Dashboard連携用
#
# 直近イベント抽出補助
# ----------------------------------------------------------

ws_events["L1"] = "Upcoming"

ws_events["L2"] = (
    '=IFERROR('
    'SORT('
    'FILTER(tblEvents[企業名],'
    'tblEvents[開始日]>=TODAY()'
    '),'
    '1,1'
    '),"")'
)


# ----------------------------------------------------------
# 見やすさ
# ----------------------------------------------------------

ws_events.sheet_view.showGridLines = True


# ==========================================================
# Part3 v2.1 終了
# ==========================================================

# ==========================================================
# Part4 v2.1
# Dashboard
# ==========================================================

# ----------------------------------------------------------
# シート基本設定
# ----------------------------------------------------------

ws_dashboard.sheet_view.showGridLines = False

# 背景色（紺ベース）
for row in range(1, 40):
    for col in range(1, 25):

        cell = ws_dashboard.cell(row=row, column=col)

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="EEF2F7"
        )

# ----------------------------------------------------------
# タイトル
# ----------------------------------------------------------

ws_dashboard["A1"] = "就活管理システム v2 Pro"

ws_dashboard["A1"].font = Font(
    bold=True,
    size=18
)

# ----------------------------------------------------------
# KPIカード生成関数
# ----------------------------------------------------------

def create_dashboard_card(
    ws,
    title_cell,
    value_cell,
    title,
    formula
):
    """
    KPIカード作成
    """

    ws[title_cell] = title

    ws[title_cell].fill = header_fill
    ws[title_cell].font = header_font
    ws[title_cell].alignment = center_alignment

    ws[value_cell] = formula

    ws[value_cell].font = card_value_font
    ws[value_cell].alignment = center_alignment

    for row in range(
        ws[title_cell].row,
        ws[value_cell].row + 1
    ):
        for col in range(
            ws[title_cell].column,
            ws[title_cell].column + 2
        ):
            ws.cell(row=row, column=col).border = card_border
            ws.cell(row=row, column=col).fill = card_fill


# ----------------------------------------------------------
# KPIカード
# ----------------------------------------------------------

create_dashboard_card(
    ws_dashboard,
    "A3",
    "A4",
    "応募企業数",
    '=COUNTA(tblCompanies[企業名])'
)

create_dashboard_card(
    ws_dashboard,
    "D3",
    "D4",
    "ES提出中",
    '=COUNTIF(tblCompanies[ステータス],"ES提出中")'
)

create_dashboard_card(
    ws_dashboard,
    "G3",
    "G4",
    "面接予定数",
    '=COUNTIFS(tblCompanies[ステータス],"*面接予定*")'
)

create_dashboard_card(
    ws_dashboard,
    "J3",
    "J4",
    "結果待ち",
    '=COUNTIF(tblCompanies[ステータス],"結果待ち")'
)

create_dashboard_card(
    ws_dashboard,
    "A7",
    "A8",
    "インターン数",
    '=COUNTIFS(tblCompanies[ステータス],"*インターン*")'
)

create_dashboard_card(
    ws_dashboard,
    "D7",
    "D8",
    "内定数",
    '=COUNTIF(tblCompanies[ステータス],"内定")'
)

create_dashboard_card(
    ws_dashboard,
    "G7",
    "G8",
    "締切3日以内",
    '=COUNTIF(tblCompanies[締切警告],"🟠3日以内")+COUNTIF(tblCompanies[締切警告],"🔴今日")'
)

create_dashboard_card(
    ws_dashboard,
    "J7",
    "J8",
    "今日のタスク",
    '=COUNTIF(tblEvents[開始日],TODAY())'
)

# ----------------------------------------------------------
# KPI列幅
# ----------------------------------------------------------

for col in [
    "A","B",
    "D","E",
    "G","H",
    "J","K"
]:
    ws_dashboard.column_dimensions[col].width = 15

# ----------------------------------------------------------
# 直近タスク
# ----------------------------------------------------------

ws_dashboard["A12"] = "📌直近タスク5件"

ws_dashboard["A12"].fill = header_fill
ws_dashboard["A12"].font = header_font

for col in range(1, 7):
    ws_dashboard.cell(
        row=12,
        column=col
    ).border = card_border

# ----------------------------------------------------------
# ヘッダ
# ----------------------------------------------------------

task_headers = [
    "企業名",
    "イベント",
    "開始日",
    "終了日",
    "開始時刻",
    "備考"
]

for idx, header in enumerate(task_headers, start=1):

    c = ws_dashboard.cell(
        row=13,
        column=idx
    )

    c.value = header
    c.fill = gray_fill
    c.font = Font(bold=True)
    c.border = card_border

# ----------------------------------------------------------
# FILTER + SORT
#
# Excel365動的配列
# ----------------------------------------------------------

ws_dashboard["A14"] = (
    '=IFERROR('
    'SORT('
    'FILTER('
    'tblEvents[[企業名]:[備考]],'
    'tblEvents[開始日]>=TODAY()'
    '),'
    '3,1'
    '),'
    '"")'
)

# ----------------------------------------------------------
# 締切アラート
# ----------------------------------------------------------

ws_dashboard["I12"] = "⚠ 締切アラート"

ws_dashboard["I12"].fill = header_fill
ws_dashboard["I12"].font = header_font

for col in range(9, 13):
    ws_dashboard.cell(
        row=12,
        column=col
    ).border = card_border

alert_headers = [
    "企業名",
    "締切",
    "警告"
]

for idx, header in enumerate(alert_headers, start=9):

    c = ws_dashboard.cell(
        row=13,
        column=idx
    )

    c.value = header
    c.fill = gray_fill
    c.font = Font(bold=True)
    c.border = card_border

ws_dashboard["I14"] = (
    '=IFERROR('
    'SORT('
    'FILTER('
    'tblCompanies[[企業名]:[締切警告]],'
    'tblCompanies[締切警告]<>""'
    '),'
    '2,1'
    '),'
    '"")'
)

# ----------------------------------------------------------
# 更新情報
# ----------------------------------------------------------

ws_dashboard["A25"] = "最終更新"

ws_dashboard["B25"] = "=TODAY()"

# ----------------------------------------------------------
# 枠線
# ----------------------------------------------------------

for row in range(14, 25):
    for col in range(1, 13):

        ws_dashboard.cell(
            row=row,
            column=col
        ).border = card_border

# ----------------------------------------------------------
# 行高
# ----------------------------------------------------------

for row in range(1, 30):
    ws_dashboard.row_dimensions[row].height = 24

# ----------------------------------------------------------
# 印刷設定
# ----------------------------------------------------------

ws_dashboard.freeze_panes = "A2"

# ==========================================================
# Part4 v2.1 終了
# ==========================================================

# ==========================================================
# Part5 v2.1
# _Calc
#
# 非表示の計算エンジン
# Dashboard
# Analytics
# Timeline
#
# の元データを集約
# ==========================================================

# ----------------------------------------------------------
# シート作成
#
# Part1で未作成の場合のみ
# ----------------------------------------------------------

if "_Calc" not in wb.sheetnames:

    ws_calc = wb.create_sheet("_Calc")

else:

    ws_calc = wb["_Calc"]


# ----------------------------------------------------------
# 非表示
# ----------------------------------------------------------

ws_calc.sheet_state = "hidden"


# ----------------------------------------------------------
# ヘッダ
# ----------------------------------------------------------

calc_headers = [
    "項目",
    "値"
]

for col_num, header in enumerate(calc_headers, start=1):

    cell = ws_calc.cell(
        row=1,
        column=col_num,
        value=header
    )

    cell.fill = header_fill
    cell.font = header_font
    cell.border = card_border
    cell.alignment = center_alignment


# ==========================================================
# KPI
# ==========================================================

kpi_items = [

    (
        "応募企業数",
        '=COUNTA(tblCompanies[企業名])'
    ),

    (
        "ES提出中",
        '=COUNTIF(tblCompanies[ステータス],"ES提出中")'
    ),

    (
        "面接予定数",
        '=COUNTIFS(tblCompanies[ステータス],"*面接予定*")'
    ),

    (
        "結果待ち",
        '=COUNTIF(tblCompanies[ステータス],"結果待ち")'
    ),

    (
        "インターン数",
        '=COUNTIFS(tblCompanies[ステータス],"*インターン*")'
    ),

    (
        "内定数",
        '=COUNTIF(tblCompanies[ステータス],"内定")'
    ),

    (
        "締切3日以内件数",
        '=COUNTIF(tblCompanies[締切警告],"🟠3日以内")+COUNTIF(tblCompanies[締切警告],"🔴今日")'
    ),

    (
        "今日のタスク件数",
        '=COUNTIF(tblEvents[開始日],TODAY())'
    )

]

start_row = 2

for idx, (title, formula) in enumerate(kpi_items):

    row = start_row + idx

    ws_calc.cell(
        row=row,
        column=1,
        value=title
    )

    ws_calc.cell(
        row=row,
        column=2,
        value=formula
    )


# ==========================================================
# Dashboard用
# 直近イベント
# ==========================================================

ws_calc["D1"] = "UpcomingEvents"

ws_calc["D2"] = (
    '=IFERROR('
    'SORT('
    'FILTER('
    'tblEvents[[企業名]:[備考]],'
    'tblEvents[開始日]>=TODAY()'
    '),'
    '3,1'
    '),'
    '"")'
)


# ==========================================================
# Dashboard用
# 締切警告
# ==========================================================

ws_calc["L1"] = "DeadlineAlerts"

ws_calc["L2"] = (
    '=IFERROR('
    'SORT('
    'FILTER('
    'tblCompanies[[企業名]:[締切警告]],'
    'tblCompanies[締切警告]<>""'
    '),'
    '2,1'
    '),'
    '"")'
)


# ==========================================================
# Analytics用
# 優先度集計
# ==========================================================

ws_calc["Q1"] = "Priority"

priority_rows = {
    "S": 2,
    "A": 3,
    "B": 4,
    "C": 5
}

for priority, row in priority_rows.items():

    ws_calc[f"Q{row}"] = priority

    ws_calc[f"R{row}"] = (
        f'=COUNTIF(tblCompanies[優先度],"'
        f'{priority}")'
    )


# ==========================================================
# Analytics用
# 志望度集計
# ==========================================================

ws_calc["T1"] = "Motivation"

motivation_values = [
    "★★★★★",
    "★★★★☆",
    "★★★☆☆",
    "★★☆☆☆",
    "★☆☆☆☆"
]

for idx, value in enumerate(
    motivation_values,
    start=2
):

    ws_calc[f"T{idx}"] = value

    ws_calc[f"U{idx}"] = (
        f'=COUNTIF(tblCompanies[志望度],"{value}")'
    )


# ==========================================================
# Analytics用
# ステータス集計
# ==========================================================

ws_calc["W1"] = "Status"

status_start_row = 2

for idx, status in enumerate(
    MASTER_STATUS,
    start=status_start_row
):

    ws_calc.cell(
        row=idx,
        column=23,
        value=status
    )

    ws_calc.cell(
        row=idx,
        column=24,
        value=(
            f'=COUNTIF('
            f'tblCompanies[ステータス],'
            f'"{status}"'
            f')'
        )
    )


# ==========================================================
# Timeline用
# 企業一覧
# ==========================================================

ws_calc["AA1"] = "CompanyList"

ws_calc["AA2"] = (
    '=SORT('
    'UNIQUE('
    'FILTER('
    'tblCompanies[企業名],'
    'tblCompanies[企業名]<>""'
    ')'
    '))'
)


# ==========================================================
# Timeline用
# イベント一覧
# ==========================================================

ws_calc["AD1"] = "EventSource"

ws_calc["AD2"] = (
    '=SORT('
    'FILTER('
    'tblEvents[[企業名]:[終了日]],'
    'tblEvents[企業名]<>""'
    ')'
    ')'
)


# ==========================================================
# 将来の拡張用
# ==========================================================

ws_calc["AH1"] = "Reserved"

ws_calc["AH2"] = "FutureUse"


# ==========================================================
# 列幅
# ==========================================================

for col in [
    "A","B",
    "D","L",
    "Q","R",
    "T","U",
    "W","X",
    "AA","AD",
    "AH"
]:
    ws_calc.column_dimensions[col].width = 20


# ==========================================================
# シート保護はしない
#
# 非表示のみ
# ==========================================================


# ==========================================================
# Part5 v2.1 終了
# ==========================================================
# ==========================================================
# Part6 v2.1
# Analytics
#
# _Calcを唯一のデータソースとする
# ==========================================================

# ----------------------------------------------------------
# 基本設定
# ----------------------------------------------------------

ws_analytics.sheet_view.showGridLines = False

# ----------------------------------------------------------
# タイトル
# ----------------------------------------------------------

ws_analytics["A1"] = "Analytics"

ws_analytics["A1"].font = Font(
    size=18,
    bold=True
)

# ----------------------------------------------------------
# 説明
# ----------------------------------------------------------

ws_analytics["A3"] = (
    "Companiesシートの内容に応じて"
)

ws_analytics["A4"] = (
    "グラフは自動更新されます"
)

# ----------------------------------------------------------
# 優先度集計領域
# ----------------------------------------------------------

ws_analytics["A7"] = "優先度"

ws_analytics["B7"] = "件数"

for cell in ["A7", "B7"]:

    ws_analytics[cell].fill = header_fill
    ws_analytics[cell].font = header_font
    ws_analytics[cell].border = card_border

for row in range(2, 6):

    analytics_row = row + 6

    ws_analytics[f"A{analytics_row}"] = (
        f"=_Calc!Q{row}"
    )

    ws_analytics[f"B{analytics_row}"] = (
        f"=_Calc!R{row}"
    )

# ----------------------------------------------------------
# 志望度集計領域
# ----------------------------------------------------------

ws_analytics["D7"] = "志望度"

ws_analytics["E7"] = "件数"

for cell in ["D7", "E7"]:

    ws_analytics[cell].fill = header_fill
    ws_analytics[cell].font = header_font
    ws_analytics[cell].border = card_border

for row in range(2, 7):

    analytics_row = row + 6

    ws_analytics[f"D{analytics_row}"] = (
        f"=_Calc!T{row}"
    )

    ws_analytics[f"E{analytics_row}"] = (
        f"=_Calc!U{row}"
    )

# ----------------------------------------------------------
# ステータス集計領域
# ----------------------------------------------------------

ws_analytics["G7"] = "ステータス"

ws_analytics["H7"] = "件数"

for cell in ["G7", "H7"]:

    ws_analytics[cell].fill = header_fill
    ws_analytics[cell].font = header_font
    ws_analytics[cell].border = card_border

for idx, status in enumerate(
    MASTER_STATUS,
    start=8
):

    calc_row = idx - 6

    ws_analytics[f"G{idx}"] = (
        f"=_Calc!W{calc_row}"
    )

    ws_analytics[f"H{idx}"] = (
        f"=_Calc!X{calc_row}"
    )

# ----------------------------------------------------------
# 列幅
# ----------------------------------------------------------

ws_analytics.column_dimensions["A"].width = 15
ws_analytics.column_dimensions["B"].width = 10

ws_analytics.column_dimensions["D"].width = 15
ws_analytics.column_dimensions["E"].width = 10

ws_analytics.column_dimensions["G"].width = 22
ws_analytics.column_dimensions["H"].width = 10

# ==========================================================
# グラフ①
# 優先度別棒グラフ
# ==========================================================

priority_chart = BarChart()

priority_chart.title = "優先度別"

priority_chart.y_axis.title = "件数"

priority_chart.x_axis.title = "優先度"

priority_data = Reference(
    ws_analytics,
    min_col=2,
    min_row=7,
    max_row=11
)

priority_category = Reference(
    ws_analytics,
    min_col=1,
    min_row=8,
    max_row=11
)

priority_chart.add_data(
    priority_data,
    titles_from_data=True
)

priority_chart.set_categories(
    priority_category
)

priority_chart.height = 8
priority_chart.width = 15

ws_analytics.add_chart(
    priority_chart,
    "J3"
)

# ==========================================================
# グラフ②
# ステータス別円グラフ
# ==========================================================

status_chart = PieChart()

status_chart.title = "ステータス別"

status_data = Reference(
    ws_analytics,
    min_col=8,
    min_row=7,
    max_row=7 + len(MASTER_STATUS)
)

status_category = Reference(
    ws_analytics,
    min_col=7,
    min_row=8,
    max_row=7 + len(MASTER_STATUS)
)

status_chart.add_data(
    status_data,
    titles_from_data=True
)

status_chart.set_categories(
    status_category
)

status_chart.height = 12
status_chart.width = 15

ws_analytics.add_chart(
    status_chart,
    "J20"
)

# ==========================================================
# グラフ③
# 志望度別棒グラフ
# ==========================================================

motivation_chart = BarChart()

motivation_chart.title = "志望度別"

motivation_chart.y_axis.title = "件数"

motivation_chart.x_axis.title = "志望度"

motivation_data = Reference(
    ws_analytics,
    min_col=5,
    min_row=7,
    max_row=12
)

motivation_category = Reference(
    ws_analytics,
    min_col=4,
    min_row=8,
    max_row=12
)

motivation_chart.add_data(
    motivation_data,
    titles_from_data=True
)

motivation_chart.set_categories(
    motivation_category
)

motivation_chart.height = 8
motivation_chart.width = 15

ws_analytics.add_chart(
    motivation_chart,
    "Z3"
)

# ----------------------------------------------------------
# 行高さ
# ----------------------------------------------------------

for row in range(1, 40):

    ws_analytics.row_dimensions[row].height = 22

# ----------------------------------------------------------
# 枠線
# ----------------------------------------------------------

for row in range(7, 32):

    for col in range(1, 9):

        ws_analytics.cell(
            row=row,
            column=col
        ).border = card_border

# ----------------------------------------------------------
# フリーズ
# ----------------------------------------------------------

ws_analytics.freeze_panes = "A7"

# ==========================================================
# Part6 v2.1 終了
# ==========================================================
# ==========================================================
# Part7 v2.1
# Timeline
# ==========================================================

from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

# ----------------------------------------------------------
# 基本設定
# ----------------------------------------------------------

ws_timeline.sheet_view.showGridLines = False

ws_timeline["A1"] = "企業名"

ws_timeline["A1"].fill = header_fill
ws_timeline["A1"].font = header_font
ws_timeline["A1"].border = card_border

# ----------------------------------------------------------
# 90日ヘッダ
# ----------------------------------------------------------

for day in range(90):

    col = day + 2

    cell = ws_timeline.cell(
        row=1,
        column=col
    )

    if day == 0:

        cell.value = "=TODAY()"

    else:

        prev_col = get_column_letter(col - 1)

        cell.value = (
            f"={prev_col}1+1"
        )

    cell.number_format = "m/d"

    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = card_border

# ----------------------------------------------------------
# 企業一覧
# ----------------------------------------------------------
#
# Timeline表示用
# 300社まで対応
#
# Python再実行不要
# Companies増加対応
#
# ----------------------------------------------------------

for row in range(2, 302):

    ws_timeline.cell(
        row=row,
        column=1
    ).value = (
        f'=IFERROR(INDEX('
        f'SORT(UNIQUE('
        f'FILTER(tblCompanies[企業名],'
        f'tblCompanies[企業名]<>"")),'
        f'ROW()-1),"")'
    )

# ----------------------------------------------------------
# 列幅
# ----------------------------------------------------------

ws_timeline.column_dimensions["A"].width = 30

for col in range(2, 92):

    letter = get_column_letter(col)

    ws_timeline.column_dimensions[
        letter
    ].width = 4

# ----------------------------------------------------------
# 行高さ
# ----------------------------------------------------------

for row in range(1, 302):

    ws_timeline.row_dimensions[row].height = 20

# ----------------------------------------------------------
# バー色
# ----------------------------------------------------------

timeline_fill = PatternFill(
    fill_type="solid",
    fgColor="4F81BD"
)

# ----------------------------------------------------------
# 条件付き書式
#
# Eventsの
# 開始日〜終了日
#
# に該当する日付だけ色付け
# ----------------------------------------------------------

timeline_range = "B2:CM301"

formula = (
    'COUNTIFS('
    'tblEvents[企業名],$A2,'
    'tblEvents[開始日],"<="&B$1,'
    'tblEvents[終了日],">="&B$1'
    ')>0'
)

rule = FormulaRule(
    formula=[formula],
    fill=timeline_fill
)

ws_timeline.conditional_formatting.add(
    timeline_range,
    rule
)

# ----------------------------------------------------------
# 枠線
# ----------------------------------------------------------

for row in range(1, 302):

    for col in range(1, 92):

        ws_timeline.cell(
            row=row,
            column=col
        ).border = card_border

# ----------------------------------------------------------
# フリーズ
# ----------------------------------------------------------

ws_timeline.freeze_panes = "B2"

# ----------------------------------------------------------
# 説明
# ----------------------------------------------------------

ws_timeline["A305"] = (
    "Eventsシートの開始日〜終了日で"
)

ws_timeline["A306"] = (
    "タイムラインが自動表示されます"
)

# ==========================================================
# 保存処理
# ==========================================================

wb.save(OUTPUT_FILE)

print(
    f"作成完了: {OUTPUT_FILE}"
)

# ==========================================================
# End
# ==========================================================