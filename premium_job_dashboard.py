import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
from collections import defaultdict
from typing import List, Dict, Any

class JobHuntingAppGenerator:
    def __init__(self, output_path: str = "就活管理システム.xlsx"):
        self.output_path = output_path
        self.wb = None
        
        # 統一カラーパレット定義
        self.NAVY_DARK = "1B365D"
        self.NAVY_LIGHT = "F0F4F8"
        self.WHITE = "FFFFFF"
        self.GRAY_TEXT = "595959"
        self.GRAY_BORDER = "D9D9D9"
        
        # 優先度別の配色設定
        self.PRIORITY_COLORS = {
            "S": {"fill": "FFD2D2", "font": "990000"},
            "A": {"fill": "FFE6CC", "font": "B35900"},
            "B": {"fill": "D2E5FF", "font": "004C99"},
            "C": {"fill": "E5E5E5", "font": "595959"}
        }
        
        # イベントガントバー用の配色定義
        self.EVENT_COLORS = {
            "ES": {"fill": "FFF2CC", "font": "7F6000"},
            "SPI": {"fill": "E2EFDA", "font": "375623"},
            "GD": {"fill": "E1D5E7", "font": "613A70"},
            "面接": {"fill": "DDEBF7", "font": "1F4E78"},
            "インターン": {"fill": "FCE4D6", "font": "C65911"},
            "内定": {"fill": "C9F2C9", "font": "1E461E"}
        }
        
        # ドロップダウン用のマスター選択肢
        self.STATUSES = [
            "応募済", "ES提出中", "ES通過", "ES不通過", "SPI受験", "SPI通過", "SPI不通過",
            "GD予定", "GD通過", "GD不通過", "一次面接通過", "二次面接通過", "最終面接通過",
            "結果待ち", "インターン参加予定", "インターン参加中", "インターン決定", "内定", "辞退"
        ]
        self.PRIORITIES = ["S", "A", "B", "C"]
        self.STARS = ["★★★★★", "★★★★☆", "★★★☆☆", "★★☆☆☆", "★☆☆☆☆"]

    def load_or_create_data(self):
        """Excelファイルが存在する場合はそれを読み込み、存在しない場合は初期サンプルデータを作成します。"""
        if os.path.exists(self.output_path):
            print(f"既存のファイル '{self.output_path}' を読み込んでタイムラインを更新します...")
            self.wb = openpyxl.load_workbook(self.output_path)
            
            # 既存の動的生成シート(Dashboard, Timeline, Analytics)を一旦削除して再構築する
            for sheet_name in ["Dashboard", "Timeline", "Analytics"]:
                if sheet_name in self.wb.sheetnames:
                    self.wb.remove(self.wb[sheet_name])
            
            # CompaniesとEventsから最新データを読み出す
            companies = self._read_companies_from_sheet()
            events = self._read_events_from_sheet()
        else:
            print(f"新規ファイルを作成し、サンプルデータを投入します...")
            self.wb = openpyxl.Workbook()
            if "Sheet" in self.wb.sheetnames:
                self.wb.remove(self.wb["Sheet"])
                
            # 初期サンプルデータ
            companies = [
                ["PwCコンサルティング", "コンサルティング", "S", "★★★★★", "面接中", "2026-05-10", "2026-06-05", "第一志望。面接対策必須"],
                ["三菱商事", "商社", "S", "★★★★★", "ES提出中", "2026-06-01", "2026-06-25", "OB訪問3名実施済"],
                ["トヨタ自動車", "メーカー", "A", "★★★★☆", "インターン参加予定", "2026-05-15", "2026-06-10", "インターン選考通過、8月参加"],
                ["リクルート", "IT・Web", "A", "★★★★☆", "SPI受験", "2026-05-20", "2026-06-15", "テストセンター受験予定"],
                ["野村総合研究所", "シンクタンク", "A", "★★★★☆", "GD予定", "2026-05-12", "2026-06-12", "グループディスカッション対策"],
                ["ソニー", "メーカー", "B", "★★★☆☆", "応募済", "2026-06-05", "2026-06-30", "本選考エントリー"],
                ["楽天グループ", "IT・Web", "B", "★★★☆☆", "内定", "2026-04-15", "2026-05-10", "早期内定。承諾保留中"],
                ["アクセンチュア", "コンサルティング", "B", "★★★☆☆", "ES通過", "2026-05-01", "2026-05-25", "次ステージ面接"],
                ["みずほフィナンシャルG", "金融", "C", "★★☆☆☆", "辞退", "2026-05-05", "2026-05-20", "他社内定のため辞退"]
            ]
            events = [
                ["PwCコンサルティング", "ES", "2026-06-01", "2026-06-05", "志望動機のブラッシュアップ"],
                ["PwCコンサルティング", "面接", "2026-06-10", "2026-06-14", "二次面接（ケース含む）"],
                ["PwCコンサルティング", "インターン", "2026-06-22", "2026-06-26", "夏期5daysインターン"],
                ["三菱商事", "ES", "2026-06-15", "2026-06-25", "インターンエントリー"],
                ["トヨタ自動車", "面接", "2026-06-05", "2026-06-09", "インターン面接（パス済）"],
                ["リクルート", "SPI", "2026-06-12", "2026-06-18", "Webテスト受験期間"],
                ["野村総合研究所", "GD", "2026-06-18", "2026-06-21", "テーマ：新規事業立案"],
                ["ソニー", "ES", "2026-06-20", "2026-06-30", "技術職エントリーシート"]
            ]
            # マスターシートの新規作成
            self._init_companies_sheet(companies)
            self._init_events_sheet(events)
            
        return companies, events

    def _read_companies_from_sheet(self) -> List[List[Any]]:
        ws = self.wb["Companies"]
        data = []
        for row in ws.iter_rows(min_row=2, max_col=8, values_only=True):
            if row[0]: # 企業名が入っている行のみ取得
                data.append(list(row))
        return data

    def _read_events_from_sheet(self) -> List[List[Any]]:
        ws = self.wb["Events"]
        data = []
        for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
            if row[0]: # 企業名が入っている行のみ取得
                data.append(list(row))
        return data

    def update_application(self):
        """全体の生成・更新ワークフローを実行します。"""
        companies, events = self.load_or_create_data()
        
        # 動的シートの作成
        self._create_dashboard_sheet()
        self._create_timeline_sheet(companies, events)
        self._create_analytics_sheet()
        
        self.wb.save(self.output_path)
        print(f"すべてのシートが正常に更新されました！保存先: {self.output_path}")

    def _apply_base_grid(self, ws):
        ws.views.sheetView[0].showGridLines = True

    def _init_companies_sheet(self, data):
        ws = self.wb.create_sheet(title="Companies")
        self._apply_base_grid(ws)
        headers = ["企業名", "業界", "優先度", "志望度", "ステータス", "応募日", "ES締切", "備考"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(name="Segoe UI", size=11, bold=True, color=self.WHITE)
            cell.fill = PatternFill(start_color=self.NAVY_DARK, end_color=self.NAVY_DARK, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        thin_border = Border(left=Side(style='thin', color=self.GRAY_BORDER), right=Side(style='thin', color=self.GRAY_BORDER),
                             top=Side(style='thin', color=self.GRAY_BORDER), bottom=Side(style='thin', color=self.GRAY_BORDER))
                             
        for r_idx, row_data in enumerate(data, start=2):
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = Font(name="Segoe UI", size=10)
                cell.border = thin_border
                
        # 右側の隠しエリアにドロップダウン用のマスタデータを書き込む
        ws.cell(row=1, column=10, value="ステータスマスタ").font = Font(bold=True)
        for i, val in enumerate(self.STATUSES, start=2): ws.cell(row=i, column=10, value=val)
        ws.cell(row=1, column=11, value="優先度マスタ").font = Font(bold=True)
        for i, val in enumerate(self.PRIORITIES, start=2): ws.cell(row=i, column=11, value=val)
        ws.cell(row=1, column=12, value="志望度マスタ").font = Font(bold=True)
        for i, val in enumerate(self.STARS, start=2): ws.cell(row=i, column=12, value=val)

        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['H'].width = 30

    def _init_events_sheet(self, data):
        ws = self.wb.create_sheet(title="Events")
        self._apply_base_grid(ws)
        headers = ["企業名", "イベント種別", "開始日", "終了日", "備考"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(name="Segoe UI", size=11, bold=True, color=self.WHITE)
            cell.fill = PatternFill(start_color=self.NAVY_DARK, end_color=self.NAVY_DARK, fill_type="solid")
            
        thin_border = Border(left=Side(style='thin', color=self.GRAY_BORDER), right=Side(style='thin', color=self.GRAY_BORDER),
                             top=Side(style='thin', color=self.GRAY_BORDER), bottom=Side(style='thin', color=self.GRAY_BORDER))
        for r_idx, row_data in enumerate(data, start=2):
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = Font(name="Segoe UI", size=10)
                cell.border = thin_border
        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['E'].width = 32

    def _create_dashboard_sheet(self):
        ws = self.wb.create_sheet(title="Dashboard", index=0)
        self._apply_base_grid(ws)
        
        ws.merge_cells("A1:N1")
        title_cell = ws["A1"]
        title_cell.value = "  就活状況ダッシュボード (Dashboard)"
        title_cell.font = Font(name="Segoe UI", size=16, bold=True, color=self.WHITE)
        title_cell.fill = PatternFill(start_color=self.NAVY_DARK, end_color=self.NAVY_DARK, fill_type="solid")
        title_cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[1].height = 40
        
        card_labels = [
            ("応募企業数", "=COUNTA(Companies!A2:A100)"),
            ("ES提出中", '=COUNTIF(Companies!E2:E100, "ES提出中")'),
            ("ES通過", '=COUNTIF(Companies!E2:E100, "ES通過")'),
            ("SPI受験", '=COUNTIF(Companies!E2:E100, "SPI受験")'),
            ("面接中", '=COUNTIF(Companies!E2:E100, "*面接*") + COUNTIF(Companies!E2:E100, "GD予定")'),
            ("結果待ち", '=COUNTIF(Companies!E2:E100, "結果待ち")'),
            ("インターン決定", '=COUNTIF(Companies!E2:E100, "インターン決定") + COUNTIF(Companies!E2:E100, "インターン参加予定")'),
            ("内定", '=COUNTIF(Companies!E2:E100, "内定")')
        ]
        
        card_positions = [
            ("B3", "B4"), ("E3", "E4"), ("H3", "H4"), ("K3", "K4"),
            ("B6", "B7"), ("E6", "E7"), ("H6", "H7"), ("K6", "K7")
        ]
        thin_border = Border(left=Side(style='thin', color=self.GRAY_BORDER), right=Side(style='thin', color=self.GRAY_BORDER),
                             top=Side(style='thin', color=self.GRAY_BORDER), bottom=Side(style='thin', color=self.GRAY_BORDER))
        
        for i, (label, formula) in enumerate(card_labels):
            pos_lbl, pos_val = card_positions[i]
            lbl_col, lbl_row = pos_lbl[0], int(pos_lbl[1:])
            val_col, val_row = pos_val[0], int(pos_val[1:])
            next_lbl_col = chr(ord(lbl_col) + 1)
            
            ws.merge_cells(f"{lbl_col}{lbl_row}:{next_lbl_col}{lbl_row}")
            ws.merge_cells(f"{val_col}{val_row}:{next_lbl_col}{val_row}")
            
            c_lbl = ws[f"{lbl_col}{lbl_row}"]
            c_lbl.value = label
            c_lbl.font = Font(name="Segoe UI", size=10, bold=True, color=self.WHITE)
            c_lbl.fill = PatternFill(start_color=self.NAVY_DARK, end_color=self.NAVY_DARK, fill_type="solid")
            c_lbl.alignment = Alignment(horizontal="center", vertical="center")
            
            c_val = ws[f"{val_col}{val_row}"]
            c_val.value = formula
            c_val.font = Font(name="Segoe UI", size=18, bold=True, color=self.NAVY_DARK)
            c_val.fill = PatternFill(start_color=self.NAVY_LIGHT, end_color=self.NAVY_LIGHT, fill_type="solid")
            c_val.alignment = Alignment(horizontal="center", vertical="center")
            
            for r in range(lbl_row, val_row + 1):
                for c in [ord(lbl_col)-64, ord(lbl_col)-63]:
                    ws.cell(row=r, column=c).border = thin_border

        # 選考通過率
        ws["B10"] = "選考通過率分析"
        ws["B10"].font = Font(name="Segoe UI", size=12, bold=True, color=self.NAVY_DARK)
        rates_data = [
            ("ES通過率", '=COUNTIF(Companies!E2:E100, "ES通過")+COUNTIF(Companies!E2:E100, "*面接*")+COUNTIF(Companies!E2:E100, "内定")', '=IF(B3>0, C12/B3, 0)'),
            ("SPI通過率", '=COUNTIF(Companies!E2:E100, "SPI通過")+COUNTIF(Companies!E2:E100, "*面接*")+COUNTIF(Companies!E2:E100, "内定")', '=IF(B3>0, C13/B3, 0)'),
            ("面接通過率", '=COUNTIF(Companies!E2:E100, "二次面接通過")+COUNTIF(Companies!E2:E100, "最終面接通過")+COUNTIF(Companies!E2:E100, "内定")', '=IF(B3>0, C14/B3, 0)'),
            ("インターン参加率", '=COUNTIF(Companies!E2:E100, "インターン参加中")+COUNTIF(Companies!E2:E100, "インターン決定")', '=IF(B3>0, C15/B3, 0)')
        ]
        for r_idx, (step, form1, form2) in enumerate(rates_data, start=12):
            ws.cell(row=r_idx, column=2, value=step)
            ws.cell(row=r_idx, column=3, value=form1).alignment = Alignment(horizontal="right")
            c2 = ws.cell(row=r_idx, column=4, value=form2)
            c2.number_format = "0.0%"
            c2.alignment = Alignment(horizontal="right")
            for c in range(2, 5): ws.cell(row=r_idx, column=c).border = thin_border

        # 優先度別集計
        ws["F10"] = "優先度別応募件数"
        ws["F10"].font = Font(name="Segoe UI", size=12, bold=True, color=self.NAVY_DARK)
        pri_data = [("Sランク", '=COUNTIF(Companies!C2:C100, "S")'), ("Aランク", '=COUNTIF(Companies!C2:C100, "A")'),
                    ("Bランク", '=COUNTIF(Companies!C2:C100, "B")'), ("Cランク", '=COUNTIF(Companies!C2:C100, "C")')]
        for r_idx, (rank, form) in enumerate(pri_data, start=12):
            ws.cell(row=r_idx, column=6, value=rank)
            ws.cell(row=r_idx, column=7, value=form).alignment = Alignment(horizontal="right")
            for c in range(6, 8): ws.cell(row=r_idx, column=c).border = thin_border

        # グラフ配置
        chart = BarChart()
        chart.type = "col"
        chart.title = "優先度別 企業数分布"
        chart.width = 11
        chart.height = 6.5
        chart.add_data(Reference(ws, min_col=7, min_row=11, max_row=15), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=6, min_row=12, max_row=15))
        chart.legend = None
        ws.add_chart(chart, "B18")

        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = 14

    def _create_timeline_sheet(self, companies: List[List[Any]], events: List[List[Any]]):
        """Timelineシートの作成: 優先度・志望度・ステータスのトリプルドロップダウン対応"""
        ws = self.wb.create_sheet(title="Timeline", index=1)
        self._apply_base_grid(ws)
        
        # 凡例表示
        ws["A1"] = "【凡例】"
        ws["A1"].font = Font(name="Segoe UI", size=10, bold=True)
        for i, (ev_type, colors) in enumerate(self.EVENT_COLORS.items(), start=2):
            cell = ws.cell(row=1, column=i, value=ev_type)
            cell.font = Font(name="Segoe UI", size=9, bold=True, color=colors["font"])
            cell.fill = PatternFill(start_color=colors["fill"], end_color=colors["fill"], fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # ヘッダー生成
        left_headers = ["企業名", "業界", "優先度", "志望度", "ステータス"]
        for i, h in enumerate(left_headers, start=1):
            ws.cell(row=3, column=i, value=h)
            ws.cell(row=3, column=i).font = Font(name="Segoe UI", size=10, bold=True, color=self.WHITE)
            ws.cell(row=3, column=i).fill = PatternFill(start_color=self.NAVY_DARK, end_color=self.NAVY_DARK, fill_type="solid")
            ws.cell(row=3, column=i).alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=3, start_column=i, end_row=4, end_column=i)
            
        # カレンダー生成 (2026年6月)
        start_date = datetime(2026, 6, 1)
        days_count = 30
        ws.merge_cells("F3:AI3")
        ws["F3"] = "2026年 6月"
        ws["F3"].font = Font(name="Segoe UI", size=11, bold=True, color=self.WHITE)
        ws["F3"].fill = PatternFill(start_color="2E4A62", end_color="2E4A62", fill_type="solid")
        ws["F3"].alignment = Alignment(horizontal="center", vertical="center")
        
        day_japanese = ["月", "火", "水", "木", "金", "土", "日"]
        for d in range(days_count):
            cur_date = start_date + timedelta(days=d)
            col_idx = 6 + d
            cell_d = ws.cell(row=4, column=col_idx)
            cell_d.value = str(cur_date.day) + "\n(" + day_japanese[cur_date.weekday()] + ")"
            cell_d.font = Font(name="Segoe UI", size=8, color=self.WHITE)
            cell_d.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_d.fill = PatternFill(start_color="4F81BD" if cur_date.weekday() == 5 else ("C0504D" if cur_date.weekday() == 6 else "415B76"), fill_type="solid")
            
        ws.freeze_panes = "F5"
        
        # イベントデータのマッピング準備
        events_by_company = defaultdict(list)
        for ev in events:
            events_by_company[ev[0]].append({
                "type": ev[1], "start": datetime.strptime(str(ev[2])[:10], "%Y-%m-%d") if isinstance(ev[2], (datetime, str)) else ev[2],
                "end": datetime.strptime(str(ev[3])[:10], "%Y-%m-%d") if isinstance(ev[3], (datetime, str)) else ev[3], "note": ev[4]
            })
            
        # 🟢 データバリデーション（ドロップダウン）の設定
        dv_priority = DataValidation(type="list", formula1="='Companies'!$K$2:$K$5", allow_blank=True)
        dv_star = DataValidation(type="list", formula1="='Companies'!$L$2:$L$6", allow_blank=True)
        dv_status = DataValidation(type="list", formula1="='Companies'!$I$2:$I$20", allow_blank=True)
        ws.add_data_validation(dv_priority)
        ws.add_data_validation(dv_star)
        ws.add_data_validation(dv_status)
        
        current_row = 5
        thin_border = Border(left=Side(style='thin', color=self.GRAY_BORDER), right=Side(style='thin', color=self.GRAY_BORDER),
                             top=Side(style='thin', color=self.GRAY_BORDER), bottom=Side(style='thin', color=self.GRAY_BORDER))
                             
        for comp in companies:
            c_name, industry, priority, star, status = comp[0], comp[1], comp[2], comp[3], comp[4]
            comp_events = events_by_company[c_name]
            
            # 多段（サブ行）判定
            allocated_rows: List[List[Dict[str, Any]]] = []
            for ev in sorted(comp_events, key=lambda x: x["start"]):
                placed = False
                for row_list in allocated_rows:
                    if all(ev["start"] > e["end"] or ev["end"] < e["start"] for e in row_list):
                        row_list.append(ev)
                        placed = True
                        break
                if not placed: allocated_rows.append([ev])
                    
            rows_needed = max(1, len(allocated_rows))
            
            # 各左側固定セルへデータ入力とマージ、およびドロップダウンの適用
            for col_idx, val in enumerate([c_name, industry, priority, star, status], start=1):
                cell_addr = f"{get_column_letter(col_idx)}{current_row}"
                target_cell = ws[cell_addr]
                target_cell.value = val
                target_cell.font = Font(name="Segoe UI", size=10)
                target_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                if rows_needed > 1:
                    ws.merge_cells(f"{cell_addr}:{get_column_letter(col_idx)}{current_row + rows_needed - 1}")
                
                # 🟢 ここでドロップダウンを Timeline シートのセルにも紐付ける
                if col_idx == 3: 
                    dv_priority.add(target_cell)
                    cfg = self.PRIORITY_COLORS.get(priority, {"fill": "FFFFFF", "font": "000000"})
                    target_cell.fill = PatternFill(start_color=cfg["fill"], end_color=cfg["fill"], fill_type="solid")
                    target_cell.font = Font(name="Segoe UI", size=10, bold=True, color=cfg["font"])
                elif col_idx == 4:
                    dv_star.add(target_cell)
                elif col_idx == 5:
                    dv_status.add(target_cell)
            
            # タイムライン背景の描画
            for r_offset in range(rows_needed):
                r_idx = current_row + r_offset
                ws.row_dimensions[r_idx].height = 22
                for d in range(days_count):
                    c_date = start_date + timedelta(days=d)
                    grid_cell = ws.cell(row=r_idx, column=6 + d)
                    grid_cell.border = thin_border
                    if c_date.weekday() == 5: grid_cell.fill = PatternFill(start_color="F2F6FB", fill_type="solid")
                    elif c_date.weekday() == 6: grid_cell.fill = PatternFill(start_color="FDF2F2", fill_type="solid")
                    
                    # 今日の日付ライン (2026-06-14を仮定)
                    if c_date.date() == datetime(2026, 6, 14).date():
                        grid_cell.border = Border(left=Side(style='medium', color="FF0000"), right=Side(style='medium', color="FF0000"),
                                                 top=Side(style='thin', color=self.GRAY_BORDER), bottom=Side(style='thin', color=self.GRAY_BORDER))
            
            # ガントバーの描画
            for r_offset, row_list in enumerate(allocated_rows):
                r_idx = current_row + r_offset
                for ev in row_list:
                    s_col = max(0, (ev["start"] - start_date).days)
                    e_col = min(days_count - 1, (ev["end"] - start_date).days)
                    start_col_idx = 6 + s_col
                    end_col_idx = 6 + e_col
                    
                    if start_col_idx <= end_col_idx:
                        ws.merge_cells(start_row=r_idx, start_column=start_col_idx, end_row=r_idx, end_column=end_col_idx)
                        bar_cell = ws.cell(row=r_idx, column=start_col_idx, value=ev["type"])
                        cfg = self.EVENT_COLORS.get(ev["type"], {"fill": "E5E5E5", "font": "000000"})
                        bar_cell.fill = PatternFill(start_color=cfg["fill"], end_color=cfg["fill"], fill_type="solid")
                        bar_cell.font = Font(name="Segoe UI", size=9, bold=True, color=cfg["font"])
                        bar_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            current_row += rows_needed
            
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 18
        for d in range(days_count): ws.column_dimensions[get_column_letter(6 + d)].width = 6

    def _create_analytics_sheet(self):
        ws = self.wb.create_sheet(title="Analytics", index=4)
        self._apply_base_grid(ws)
        ws.merge_cells("A1:K1")
        ws["A1"] = "  就職活動 選考・統計分析レポーティング (Analytics)"
        ws["A1"].font = Font(name="Segoe UI", size=14, bold=True, color=self.WHITE)
        ws["A1"].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        ws.row_dimensions[1].height = 35

        ws["A3"] = "ステータス分布データ"
        ws["A4"] = "ステータス"
        ws["B4"] = "社数"
        
        statuses = ["内定", "面接中", "ES提出中", "応募済", "インターン決定", "辞退"]
        formulas = ['=COUNTIF(Companies!E2:E100, "内定")', '=COUNTIF(Companies!E2:E100, "*面接*")',
                    '=COUNTIF(Companies!E2:E100, "ES提出中")', '=COUNTIF(Companies!E2:E100, "応募済")',
                    '=COUNTIF(Companies!E2:E100, "インターン決定")', '=COUNTIF(Companies!E2:E100, "辞退")']
        
        thin_border = Border(left=Side(style='thin', color=self.GRAY_BORDER), right=Side(style='thin', color=self.GRAY_BORDER),
                             top=Side(style='thin', color=self.GRAY_BORDER), bottom=Side(style='thin', color=self.GRAY_BORDER))
        for idx, (st, form) in enumerate(zip(statuses, formulas), start=5):
            ws.cell(row=idx, column=1, value=st)
            ws.cell(row=idx, column=2, value=form).alignment = Alignment(horizontal="right")
            ws.cell(row=idx, column=1).border = thin_border
            ws.cell(row=idx, column=2).border = thin_border

        pie = PieChart()
        pie.title = "選考ステータス比率"
        pie.add_data(Reference(ws, min_col=2, min_row=4, max_row=10), titles_from_data=True)
        pie.set_categories(Reference(ws, min_col=1, min_row=5, max_row=10))
        pie.width = 12
        pie.height = 7.5
        ws.add_chart(pie, "D3")
        ws.column_dimensions['A'].width = 16

if __name__ == "__main__":
    app = JobHuntingAppGenerator()
    app.update_application()