# ==========================================================
# job_hunting_system_v3_1.py
#
# 就活管理システム v3.1 Pro
#
# Pythonは初回生成のみ
# 以後はExcel365のみで運用
#
# openpyxl安定版
# ==========================================================

import os
from datetime import datetime

from openpyxl import Workbook

from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)

from openpyxl.worksheet.table import (
    Table,
    TableStyleInfo
)

from openpyxl.worksheet.datavalidation import (
    DataValidation
)

from openpyxl.chart import (
    BarChart,
    PieChart,
    Reference
)

from openpyxl.formatting.rule import (
    FormulaRule
)

from openpyxl.utils import (
    get_column_letter
)

# ==========================================================
# ファイル設定
# ==========================================================

OUTPUT_FILE = "就活管理システム_v3_1.xlsx"

# ==========================================================
# 上書き禁止
# ==========================================================

if os.path.exists(OUTPUT_FILE):

    print(
        f"エラー: {OUTPUT_FILE} は既に存在します。"
    )

    raise SystemExit(1)

# ==========================================================
# Workbook作成
# ==========================================================

wb = Workbook()

# ==========================================================
# シート作成
# ==========================================================

ws_dashboard = wb.active
ws_dashboard.title = "Dashboard"

ws_companies = wb.create_sheet(
    "Companies"
)

ws_events = wb.create_sheet(
    "Events"
)

ws_timeline = wb.create_sheet(
    "Timeline"
)

ws_analytics = wb.create_sheet(
    "Analytics"
)

ws_settings = wb.create_sheet(
    "Settings"
)

ws_master = wb.create_sheet(
    "_Master"
)

# ==========================================================
# カラーパレット
# ==========================================================

COLOR_NAVY = "1F3A5F"
COLOR_LIGHT_GRAY = "EAEAEA"
COLOR_WHITE = "FFFFFF"
COLOR_BORDER = "CFCFCF"

# ==========================================================
# 共通スタイル
# ==========================================================

header_fill = PatternFill(
    fill_type="solid",
    fgColor=COLOR_NAVY
)

light_fill = PatternFill(
    fill_type="solid",
    fgColor=COLOR_LIGHT_GRAY
)

header_font = Font(
    color=COLOR_WHITE,
    bold=True
)

normal_font = Font(
    color="000000"
)

thin_side = Side(
    border_style="thin",
    color=COLOR_BORDER
)

card_border = Border(
    left=thin_side,
    right=thin_side,
    top=thin_side,
    bottom=thin_side
)

center_alignment = Alignment(
    horizontal="center",
    vertical="center"
)

left_alignment = Alignment(
    horizontal="left",
    vertical="center"
)

# ==========================================================
# マスタデータ
# ==========================================================

MASTER_INDUSTRIES = [

    "コンサル",
    "商社",
    "メーカー",
    "IT・Web",
    "シンクタンク",
    "金融",
    "人材",
    "通信",
    "不動産",
    "広告",
    "インフラ",
    "物流",
    "公務員",
    "その他"

]

MASTER_PRIORITIES = [

    "S",
    "A",
    "B",
    "C"

]

MASTER_MOTIVATIONS = [

    "★★★★★",
    "★★★★☆",
    "★★★☆☆",
    "★★☆☆☆",
    "★☆☆☆☆"

]

MASTER_STATUS = [

    "応募予定",
    "応募済",

    "ES提出中",
    "ES通過",
    "ES不通過",

    "SPI受験",
    "SPI通過",
    "SPI不通過",

    "GD予定",
    "GD通過",
    "GD不通過",

    "一次面接予定",
    "一次面接通過",

    "二次面接予定",
    "二次面接通過",

    "最終面接予定",
    "最終面接通過",

    "結果待ち",

    "インターン参加予定",
    "インターン参加中",
    "インターン決定",

    "内定",

    "辞退"

]

MASTER_EVENTS = [

    "説明会",
    "ES",
    "SPI",
    "GD",
    "一次面接",
    "二次面接",
    "最終面接",
    "インターン",
    "内定"

]

# ==========================================================
# 共通関数
# ==========================================================

def style_header_row(
    worksheet,
    row_number,
    max_col
):
    """
    ヘッダー行へ共通スタイル適用
    """

    for col in range(
        1,
        max_col + 1
    ):

        cell = worksheet.cell(
            row=row_number,
            column=col
        )

        cell.fill = header_fill
        cell.font = header_font
        cell.border = card_border
        cell.alignment = center_alignment


def set_column_widths(
    worksheet,
    width_map
):
    """
    列幅設定
    """

    for col_letter, width in width_map.items():

        worksheet.column_dimensions[
            col_letter
        ].width = width


def apply_border_area(
    worksheet,
    start_row,
    end_row,
    start_col,
    end_col
):
    """
    指定範囲へ罫線適用
    """

    for row in range(
        start_row,
        end_row + 1
    ):

        for col in range(
            start_col,
            end_col + 1
        ):

            worksheet.cell(
                row=row,
                column=col
            ).border = card_border


# ==========================================================
# 現在日時
# ==========================================================

CURRENT_DATE = datetime.now()

CURRENT_DATE_TEXT = CURRENT_DATE.strftime(
    "%Y-%m-%d"
)

# ==========================================================
# Part1 End
# ==========================================================

# ==========================================================
# Part2
# _Master
# Companies
# Events
# ==========================================================

# ==========================================================
# _Master
# ==========================================================

master_headers = [
    "業界",
    "優先度",
    "志望度",
    "ステータス",
    "イベント"
]

for col, header in enumerate(master_headers, start=1):

    ws_master.cell(
        row=1,
        column=col,
        value=header
    )

style_header_row(
    ws_master,
    1,
    5
)

# ----------------------------------------------------------
# 業界
# ----------------------------------------------------------

for row, value in enumerate(
    MASTER_INDUSTRIES,
    start=2
):
    ws_master.cell(
        row=row,
        column=1,
        value=value
    )

# ----------------------------------------------------------
# 優先度
# ----------------------------------------------------------

for row, value in enumerate(
    MASTER_PRIORITIES,
    start=2
):
    ws_master.cell(
        row=row,
        column=2,
        value=value
    )

# ----------------------------------------------------------
# 志望度
# ----------------------------------------------------------

for row, value in enumerate(
    MASTER_MOTIVATIONS,
    start=2
):
    ws_master.cell(
        row=row,
        column=3,
        value=value
    )

# ----------------------------------------------------------
# ステータス
# ----------------------------------------------------------

for row, value in enumerate(
    MASTER_STATUS,
    start=2
):
    ws_master.cell(
        row=row,
        column=4,
        value=value
    )

# ----------------------------------------------------------
# イベント
# ----------------------------------------------------------

for row, value in enumerate(
    MASTER_EVENTS,
    start=2
):
    ws_master.cell(
        row=row,
        column=5,
        value=value
    )

# ----------------------------------------------------------
# 列幅
# ----------------------------------------------------------

set_column_widths(
    ws_master,
    {
        "A": 20,
        "B": 12,
        "C": 15,
        "D": 25,
        "E": 15
    }
)

# ----------------------------------------------------------
# 非表示
# ----------------------------------------------------------

ws_master.sheet_state = "hidden"

# ==========================================================
# Companies
# ==========================================================

company_headers = [

    "企業名",
    "業界",
    "優先度",
    "志望度",
    "ステータス",
    "応募日",
    "ES締切",
    "締切警告",
    "マイページURL",
    "備考"

]

for col, header in enumerate(
    company_headers,
    start=1
):

    ws_companies.cell(
        row=1,
        column=col,
        value=header
    )

style_header_row(
    ws_companies,
    1,
    len(company_headers)
)

# ----------------------------------------------------------
# 列幅
# ----------------------------------------------------------

set_column_widths(
    ws_companies,
    {
        "A": 30,
        "B": 18,
        "C": 10,
        "D": 12,
        "E": 22,
        "F": 14,
        "G": 14,
        "H": 14,
        "I": 40,
        "J": 40
    }
)

# ----------------------------------------------------------
# 日付列
# ----------------------------------------------------------

for row in range(2, 5001):

    ws_companies.cell(
        row=row,
        column=6
    ).number_format = "yyyy/mm/dd"

    ws_companies.cell(
        row=row,
        column=7
    ).number_format = "yyyy/mm/dd"

# ----------------------------------------------------------
# 締切警告
# ----------------------------------------------------------

for row in range(2, 5001):

    formula = (
        f'=IF(G{row}="","",'
        f'IF(G{row}<TODAY(),"期限切れ",'
        f'IF(G{row}=TODAY(),"今日",'
        f'IF(G{row}<=TODAY()+3,"3日以内",'
        f'IF(G{row}<=TODAY()+7,"7日以内","")))))'
    )

    ws_companies.cell(
        row=row,
        column=8,
        value=formula
    )

# ----------------------------------------------------------
# テーブル
# ----------------------------------------------------------
#
# 最小サイズで生成
# Excelが自動拡張
#
# ----------------------------------------------------------

company_table = Table(
    displayName="tblCompanies",
    ref="A1:J2"
)

company_style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)

company_table.tableStyleInfo = company_style

ws_companies.add_table(
    company_table
)

# ----------------------------------------------------------
# フリーズ
# ----------------------------------------------------------

ws_companies.freeze_panes = "A2"

# ==========================================================
# Events
# ==========================================================

event_headers = [

    "企業名",
    "イベント",
    "開始日",
    "終了日",
    "開始時刻",
    "備考"

]

for col, header in enumerate(
    event_headers,
    start=1
):

    ws_events.cell(
        row=1,
        column=col,
        value=header
    )

style_header_row(
    ws_events,
    1,
    len(event_headers)
)

# ----------------------------------------------------------
# 列幅
# ----------------------------------------------------------

set_column_widths(
    ws_events,
    {
        "A": 30,
        "B": 15,
        "C": 14,
        "D": 14,
        "E": 12,
        "F": 40
    }
)

# ----------------------------------------------------------
# 日付書式
# ----------------------------------------------------------

for row in range(2, 5001):

    ws_events.cell(
        row=row,
        column=3
    ).number_format = "yyyy/mm/dd"

    ws_events.cell(
        row=row,
        column=4
    ).number_format = "yyyy/mm/dd"

# ----------------------------------------------------------
# 時刻書式
# ----------------------------------------------------------

for row in range(2, 5001):

    ws_events.cell(
        row=row,
        column=5
    ).number_format = "hh:mm"

# ----------------------------------------------------------
# テーブル
# ----------------------------------------------------------

event_table = Table(
    displayName="tblEvents",
    ref="A1:F2"
)

event_style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)

event_table.tableStyleInfo = event_style

ws_events.add_table(
    event_table
)

# ----------------------------------------------------------
# フリーズ
# ----------------------------------------------------------

ws_events.freeze_panes = "A2"

# ==========================================================
# Part2 End
# ==========================================================

# ==========================================================
# Part3
# Dashboard
# ==========================================================

# ----------------------------------------------------------
# 基本設定
# ----------------------------------------------------------

ws_dashboard.sheet_view.showGridLines = False

# ----------------------------------------------------------
# タイトル
# ----------------------------------------------------------

ws_dashboard["A1"] = "就活管理システム v3.1"

ws_dashboard["A1"].font = Font(
    size=18,
    bold=True,
    color="000000"
)

# ----------------------------------------------------------
# Dashboard列幅
# ----------------------------------------------------------

set_column_widths(
    ws_dashboard,
    {
        "A": 16,
        "B": 16,
        "C": 4,
        "D": 16,
        "E": 16,
        "F": 4,
        "G": 16,
        "H": 16,
        "I": 4,
        "J": 16,
        "K": 16
    }
)

# ==========================================================
# KPIカード共通関数
# ==========================================================

def create_kpi_card(
    worksheet,
    title_cell,
    value_cell,
    title,
    formula
):

    worksheet[title_cell] = title
    worksheet[value_cell] = formula

    worksheet[title_cell].fill = header_fill
    worksheet[title_cell].font = header_font
    worksheet[title_cell].alignment = center_alignment
    worksheet[title_cell].border = card_border

    worksheet[value_cell].fill = light_fill
    worksheet[value_cell].alignment = center_alignment
    worksheet[value_cell].border = card_border

    worksheet[value_cell].font = Font(
        size=16,
        bold=True
    )

# ==========================================================
# KPI
# ==========================================================

# ----------------------------------------------------------
# 応募企業数
# ----------------------------------------------------------

create_kpi_card(
    ws_dashboard,
    "A3",
    "A4",
    "応募企業数",
    '=COUNTA(Companies!A:A)-1'
)

# ----------------------------------------------------------
# ES提出中
# ----------------------------------------------------------

create_kpi_card(
    ws_dashboard,
    "D3",
    "D4",
    "ES提出中",
    '=COUNTIF(Companies!E:E,"ES提出中")'
)

# ----------------------------------------------------------
# 面接予定数
# ----------------------------------------------------------

create_kpi_card(
    ws_dashboard,
    "G3",
    "G4",
    "面接予定数",
    '=COUNTIF(Companies!E:E,"一次面接予定")+'
    'COUNTIF(Companies!E:E,"二次面接予定")+'
    'COUNTIF(Companies!E:E,"最終面接予定")'
)

# ----------------------------------------------------------
# 結果待ち
# ----------------------------------------------------------

create_kpi_card(
    ws_dashboard,
    "J3",
    "J4",
    "結果待ち",
    '=COUNTIF(Companies!E:E,"結果待ち")'
)

# ----------------------------------------------------------
# インターン
# ----------------------------------------------------------

create_kpi_card(
    ws_dashboard,
    "A7",
    "A8",
    "インターン",
    '=COUNTIF(Companies!E:E,"インターン参加予定")+'
    'COUNTIF(Companies!E:E,"インターン参加中")+'
    'COUNTIF(Companies!E:E,"インターン決定")'
)

# ----------------------------------------------------------
# 内定
# ----------------------------------------------------------

create_kpi_card(
    ws_dashboard,
    "D7",
    "D8",
    "内定",
    '=COUNTIF(Companies!E:E,"内定")'
)

# ----------------------------------------------------------
# 締切3日以内
# ----------------------------------------------------------

create_kpi_card(
    ws_dashboard,
    "G7",
    "G8",
    "締切3日以内",
    '=COUNTIF(Companies!H:H,"今日")+'
    'COUNTIF(Companies!H:H,"3日以内")'
)

# ----------------------------------------------------------
# 今日のタスク
# ----------------------------------------------------------

create_kpi_card(
    ws_dashboard,
    "J7",
    "J8",
    "今日のタスク",
    '=COUNTIF(Events!C:C,TODAY())'
)

# ==========================================================
# 直近イベント表示エリア
# ==========================================================

ws_dashboard["A12"] = "直近イベント"

ws_dashboard["A12"].fill = header_fill
ws_dashboard["A12"].font = header_font
ws_dashboard["A12"].border = card_border

# ヘッダ

event_headers = [
    "企業名",
    "イベント",
    "開始日",
    "終了日"
]

for col, header in enumerate(
    event_headers,
    start=1
):

    cell = ws_dashboard.cell(
        row=13,
        column=col,
        value=header
    )

    cell.fill = light_fill
    cell.border = card_border
    cell.alignment = center_alignment

# ----------------------------------------------------------
# 直近10件表示
# ----------------------------------------------------------

for row in range(14, 24):

    source_row = row - 12

    ws_dashboard.cell(
        row=row,
        column=1,
        value=f"=Events!A{source_row}"
    )

    ws_dashboard.cell(
        row=row,
        column=2,
        value=f"=Events!B{source_row}"
    )

    ws_dashboard.cell(
        row=row,
        column=3,
        value=f"=Events!C{source_row}"
    )

    ws_dashboard.cell(
        row=row,
        column=4,
        value=f"=Events!D{source_row}"
    )

# ----------------------------------------------------------
# 枠線
# ----------------------------------------------------------

apply_border_area(
    ws_dashboard,
    13,
    23,
    1,
    4
)

# ==========================================================
# 締切一覧
# ==========================================================

ws_dashboard["G12"] = "締切一覧"

ws_dashboard["G12"].fill = header_fill
ws_dashboard["G12"].font = header_font
ws_dashboard["G12"].border = card_border

deadline_headers = [
    "企業名",
    "ES締切",
    "警告"
]

for idx, header in enumerate(
    deadline_headers,
    start=7
):

    cell = ws_dashboard.cell(
        row=13,
        column=idx,
        value=header
    )

    cell.fill = light_fill
    cell.border = card_border
    cell.alignment = center_alignment

# ----------------------------------------------------------
# 上位10件
# ----------------------------------------------------------

for row in range(14, 24):

    source_row = row - 12

    ws_dashboard.cell(
        row=row,
        column=7,
        value=f"=Companies!A{source_row}"
    )

    ws_dashboard.cell(
        row=row,
        column=8,
        value=f"=Companies!G{source_row}"
    )

    ws_dashboard.cell(
        row=row,
        column=9,
        value=f"=Companies!H{source_row}"
    )

apply_border_area(
    ws_dashboard,
    13,
    23,
    7,
    9
)

# ==========================================================
# 行高さ
# ==========================================================

for row in range(1, 40):

    ws_dashboard.row_dimensions[
        row
    ].height = 24

# ==========================================================
# フリーズ
# ==========================================================

ws_dashboard.freeze_panes = "A3"

# ==========================================================
# Part3 End
# ==========================================================

# ==========================================================
# Part4
# Analytics
# ==========================================================

# ----------------------------------------------------------
# 基本設定
# ----------------------------------------------------------

ws_analytics.sheet_view.showGridLines = False

ws_analytics["A1"] = "Analytics"

ws_analytics["A1"].font = Font(
    size=18,
    bold=True
)

# ==========================================================
# 優先度分析
# ==========================================================

ws_analytics["A3"] = "優先度"
ws_analytics["B3"] = "件数"

style_header_row(
    ws_analytics,
    3,
    2
)

# ----------------------------------------------------------
# マスタ参照
# ----------------------------------------------------------

for row in range(2, 102):

    analytics_row = row + 2

    ws_analytics.cell(
        row=analytics_row,
        column=1,
        value=f"=_Master!B{row}"
    )

    ws_analytics.cell(
        row=analytics_row,
        column=2,
        value=(
            f'=IF(A{analytics_row}="","",'
            f'COUNTIF(Companies!C:C,A{analytics_row}))'
        )
    )

# ==========================================================
# 志望度分析
# ==========================================================

ws_analytics["D3"] = "志望度"
ws_analytics["E3"] = "件数"

for cell in ["D3", "E3"]:

    ws_analytics[cell].fill = header_fill
    ws_analytics[cell].font = header_font
    ws_analytics[cell].border = card_border
    ws_analytics[cell].alignment = center_alignment

for row in range(2, 102):

    analytics_row = row + 2

    ws_analytics.cell(
        row=analytics_row,
        column=4,
        value=f"=_Master!C{row}"
    )

    ws_analytics.cell(
        row=analytics_row,
        column=5,
        value=(
            f'=IF(D{analytics_row}="","",'
            f'COUNTIF(Companies!D:D,D{analytics_row}))'
        )
    )

# ==========================================================
# ステータス分析
# ==========================================================

ws_analytics["G3"] = "ステータス"
ws_analytics["H3"] = "件数"

for cell in ["G3", "H3"]:

    ws_analytics[cell].fill = header_fill
    ws_analytics[cell].font = header_font
    ws_analytics[cell].border = card_border
    ws_analytics[cell].alignment = center_alignment

for row in range(2, 202):

    analytics_row = row + 2

    ws_analytics.cell(
        row=analytics_row,
        column=7,
        value=f"=_Master!D{row}"
    )

    ws_analytics.cell(
        row=analytics_row,
        column=8,
        value=(
            f'=IF(G{analytics_row}="","",'
            f'COUNTIF(Companies!E:E,G{analytics_row}))'
        )
    )

# ==========================================================
# 列幅
# ==========================================================

set_column_widths(
    ws_analytics,
    {
        "A": 15,
        "B": 10,
        "D": 15,
        "E": 10,
        "G": 25,
        "H": 10
    }
)

# ==========================================================
# 枠線
# ==========================================================

apply_border_area(
    ws_analytics,
    3,
    105,
    1,
    2
)

apply_border_area(
    ws_analytics,
    3,
    105,
    4,
    5
)

apply_border_area(
    ws_analytics,
    3,
    205,
    7,
    8
)

# ==========================================================
# グラフ①
# 優先度別棒グラフ
# ==========================================================

priority_chart = BarChart()

priority_chart.title = "優先度別"

priority_chart.y_axis.title = "件数"

priority_chart.x_axis.title = "優先度"

priority_data = Reference(
    ws_analytics,
    min_col=2,
    min_row=3,
    max_row=10
)

priority_categories = Reference(
    ws_analytics,
    min_col=1,
    min_row=4,
    max_row=10
)

priority_chart.add_data(
    priority_data,
    titles_from_data=True
)

priority_chart.set_categories(
    priority_categories
)

priority_chart.height = 8
priority_chart.width = 15

ws_analytics.add_chart(
    priority_chart,
    "J3"
)

# ==========================================================
# グラフ②
# 志望度別棒グラフ
# ==========================================================

motivation_chart = BarChart()

motivation_chart.title = "志望度別"

motivation_chart.y_axis.title = "件数"

motivation_chart.x_axis.title = "志望度"

motivation_data = Reference(
    ws_analytics,
    min_col=5,
    min_row=3,
    max_row=10
)

motivation_categories = Reference(
    ws_analytics,
    min_col=4,
    min_row=4,
    max_row=10
)

motivation_chart.add_data(
    motivation_data,
    titles_from_data=True
)

motivation_chart.set_categories(
    motivation_categories
)

motivation_chart.height = 8
motivation_chart.width = 15

ws_analytics.add_chart(
    motivation_chart,
    "J20"
)

# ==========================================================
# グラフ③
# ステータス別円グラフ
# ==========================================================

status_chart = PieChart()

status_chart.title = "ステータス別"

status_data = Reference(
    ws_analytics,
    min_col=8,
    min_row=3,
    max_row=30
)

status_categories = Reference(
    ws_analytics,
    min_col=7,
    min_row=4,
    max_row=30
)

status_chart.add_data(
    status_data,
    titles_from_data=True
)

status_chart.set_categories(
    status_categories
)

status_chart.height = 12
status_chart.width = 15

ws_analytics.add_chart(
    status_chart,
    "Z3"
)

# ==========================================================
# フリーズ
# ==========================================================

ws_analytics.freeze_panes = "A4"

# ==========================================================
# 行高さ
# ==========================================================

for row in range(1, 220):

    ws_analytics.row_dimensions[
        row
    ].height = 22

# ==========================================================
# Part4 End
# ==========================================================

# ==========================================================
# Part5
# Timeline
# ==========================================================

# ----------------------------------------------------------
# 基本設定
# ----------------------------------------------------------

ws_timeline.sheet_view.showGridLines = False

# ----------------------------------------------------------
# タイトル
# ----------------------------------------------------------

ws_timeline["A1"] = "企業名"
ws_timeline["B1"] = "イベント"
ws_timeline["C1"] = "開始日"
ws_timeline["D1"] = "終了日"

style_header_row(
    ws_timeline,
    1,
    4
)

# ==========================================================
# Timelineヘッダ
# ==========================================================

START_TIMELINE_COL = 5  # E列

for day in range(90):

    col = START_TIMELINE_COL + day

    cell = ws_timeline.cell(
        row=1,
        column=col
    )

    if day == 0:

        cell.value = "=TODAY()"

    else:

        prev_col = get_column_letter(col - 1)

        cell.value = (
            f"={prev_col}1+1"
        )

    cell.number_format = "m/d"

    cell.fill = header_fill
    cell.font = header_font
    cell.border = card_border
    cell.alignment = center_alignment

# ==========================================================
# Events参照
# ==========================================================

MAX_TIMELINE_ROWS = 300

for row in range(
    2,
    MAX_TIMELINE_ROWS + 2
):

    event_row = row

    # 企業名

    ws_timeline.cell(
        row=row,
        column=1,
        value=f"=Events!A{event_row}"
    )

    # イベント

    ws_timeline.cell(
        row=row,
        column=2,
        value=f"=Events!B{event_row}"
    )

    # 開始日

    ws_timeline.cell(
        row=row,
        column=3,
        value=f"=Events!C{event_row}"
    )

    # 終了日

    ws_timeline.cell(
        row=row,
        column=4,
        value=f"=Events!D{event_row}"
    )

# ==========================================================
# 列幅
# ==========================================================

set_column_widths(
    ws_timeline,
    {
        "A": 28,
        "B": 16,
        "C": 12,
        "D": 12
    }
)

for col in range(
    START_TIMELINE_COL,
    START_TIMELINE_COL + 90
):

    ws_timeline.column_dimensions[
        get_column_letter(col)
    ].width = 4

# ==========================================================
# 日付書式
# ==========================================================

for row in range(
    2,
    MAX_TIMELINE_ROWS + 2
):

    ws_timeline.cell(
        row=row,
        column=3
    ).number_format = "yyyy/mm/dd"

    ws_timeline.cell(
        row=row,
        column=4
    ).number_format = "yyyy/mm/dd"

# ==========================================================
# 条件付き書式
# ==========================================================

timeline_fill = PatternFill(
    fill_type="solid",
    fgColor=COLOR_NAVY
)

timeline_start_col_letter = get_column_letter(
    START_TIMELINE_COL
)

timeline_end_col_letter = get_column_letter(
    START_TIMELINE_COL + 89
)

timeline_range = (
    f"{timeline_start_col_letter}2:"
    f"{timeline_end_col_letter}"
    f"{MAX_TIMELINE_ROWS + 1}"
)

rule_formula = (
    f"=AND("
    f"{timeline_start_col_letter}$1>=$C2,"
    f"{timeline_start_col_letter}$1<=$D2"
    f")"
)

timeline_rule = FormulaRule(
    formula=[rule_formula],
    fill=timeline_fill
)

ws_timeline.conditional_formatting.add(
    timeline_range,
    timeline_rule
)

# ==========================================================
# 枠線
# ==========================================================

apply_border_area(
    ws_timeline,
    1,
    MAX_TIMELINE_ROWS + 1,
    1,
    START_TIMELINE_COL + 89
)

# ==========================================================
# 凡例
# ==========================================================

legend_row = MAX_TIMELINE_ROWS + 4

ws_timeline.cell(
    row=legend_row,
    column=1,
    value="凡例"
)

ws_timeline.cell(
    row=legend_row,
    column=2,
    value="イベント期間"
)

ws_timeline.cell(
    row=legend_row,
    column=2
).fill = timeline_fill

# ==========================================================
# フリーズ
# ==========================================================

ws_timeline.freeze_panes = "E2"

# ==========================================================
# 行高さ
# ==========================================================

for row in range(
    1,
    MAX_TIMELINE_ROWS + 10
):

    ws_timeline.row_dimensions[
        row
    ].height = 20

# ==========================================================
# Part5 End
# ==========================================================

# ==========================================================
# Part6
# Settings
# DataValidation
# Save
# ==========================================================

# ==========================================================
# Settings
# ==========================================================

ws_settings["A1"] = "項目"
ws_settings["B1"] = "値"

style_header_row(
    ws_settings,
    1,
    2
)

ws_settings["A2"] = "Version"
ws_settings["B2"] = "v3.1"

ws_settings["A3"] = "作成日"
ws_settings["B3"] = CURRENT_DATE_TEXT

ws_settings["A4"] = "最終更新日"
ws_settings["B4"] = CURRENT_DATE_TEXT

set_column_widths(
    ws_settings,
    {
        "A": 20,
        "B": 30
    }
)

apply_border_area(
    ws_settings,
    1,
    4,
    1,
    2
)

# ==========================================================
# Data Validation
# ==========================================================

MAX_INPUT_ROW = 5000

# ----------------------------------------------------------
# 業界
# ----------------------------------------------------------

dv_industry = DataValidation(
    type="list",
    formula1="=_Master!$A$2:$A$100",
    allow_blank=True
)

# ----------------------------------------------------------
# 優先度
# ----------------------------------------------------------

dv_priority = DataValidation(
    type="list",
    formula1="=_Master!$B$2:$B$100",
    allow_blank=True
)

# ----------------------------------------------------------
# 志望度
# ----------------------------------------------------------

dv_motivation = DataValidation(
    type="list",
    formula1="=_Master!$C$2:$C$100",
    allow_blank=True
)

# ----------------------------------------------------------
# ステータス
# ----------------------------------------------------------

dv_status = DataValidation(
    type="list",
    formula1="=_Master!$D$2:$D$300",
    allow_blank=True
)

# ----------------------------------------------------------
# イベント
# ----------------------------------------------------------

dv_event = DataValidation(
    type="list",
    formula1="=_Master!$E$2:$E$100",
    allow_blank=True
)

# ==========================================================
# 登録
# ==========================================================

ws_companies.add_data_validation(
    dv_industry
)

ws_companies.add_data_validation(
    dv_priority
)

ws_companies.add_data_validation(
    dv_motivation
)

ws_companies.add_data_validation(
    dv_status
)

ws_events.add_data_validation(
    dv_event
)

# ==========================================================
# Companies
# ==========================================================

dv_industry.add(
    f"B2:B{MAX_INPUT_ROW}"
)

dv_priority.add(
    f"C2:C{MAX_INPUT_ROW}"
)

dv_motivation.add(
    f"D2:D{MAX_INPUT_ROW}"
)

dv_status.add(
    f"E2:E{MAX_INPUT_ROW}"
)

# ==========================================================
# Events
# ==========================================================

dv_event.add(
    f"B2:B{MAX_INPUT_ROW}"
)

# ==========================================================
# 締切警告色付け
# ==========================================================

warning_red = PatternFill(
    fill_type="solid",
    fgColor="FF9999"
)

warning_orange = PatternFill(
    fill_type="solid",
    fgColor="FFCC99"
)

warning_yellow = PatternFill(
    fill_type="solid",
    fgColor="FFFF99"
)

# 今日

ws_companies.conditional_formatting.add(
    "H2:H5000",
    FormulaRule(
        formula=['H2="今日"'],
        fill=warning_red
    )
)

# 期限切れ

ws_companies.conditional_formatting.add(
    "H2:H5000",
    FormulaRule(
        formula=['H2="期限切れ"'],
        fill=warning_red
    )
)

# 3日以内

ws_companies.conditional_formatting.add(
    "H2:H5000",
    FormulaRule(
        formula=['H2="3日以内"'],
        fill=warning_orange
    )
)

# 7日以内

ws_companies.conditional_formatting.add(
    "H2:H5000",
    FormulaRule(
        formula=['H2="7日以内"'],
        fill=warning_yellow
    )
)

# ==========================================================
# シート保護
# ==========================================================

for ws in [

    ws_dashboard,
    ws_timeline,
    ws_analytics,
    ws_settings

]:

    ws.protection.sheet = True

# ==========================================================
# 表示位置
# ==========================================================

wb.active = 0

# ==========================================================
# 保存
# ==========================================================

wb.save(
    OUTPUT_FILE
)

print(
    f"作成完了: {OUTPUT_FILE}"
)

# ==========================================================
# End
# ==========================================================
